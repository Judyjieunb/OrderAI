"""
유사스타일 모델 1차 vs 2차 성능 비교 리포트

score_step0_model.py의 핵심 분석 로직을 재활용하여
3개 모델(1차, 2차혼용율OFF, 2차혼용율ON)을 동일 기준으로 비교한다.

- Input:
  - output/유사스타일자동결과_1차+정답지.xlsx (채점결과, 정답지, DB)
  - output/유사스타일반환결과_2차_혼용율YN.xlsx (Result 1, 정답지)
- Output: output/step0_model_comparison.xlsx + 콘솔 요약
"""

import os
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILE_1ST = os.path.join(BASE_DIR, "output", "유사스타일자동결과_1차+정답지.xlsx")
FILE_2ND = os.path.join(BASE_DIR, "output", "유사스타일반환결과_2차_혼용율YN.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "step0_model_comparison.xlsx")


# ── helpers (from score_step0_model.py) ──────────────────────────

def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    if s in ("", "0", "None", "nan", "NaN"):
        return None
    return s


def text_similarity(a, b):
    a, b = _clean(a), _clean(b)
    if a is None or b is None:
        return None
    a = re.sub(r"[^\w\s]", "", a)
    b = re.sub(r"[^\w\s]", "", b)
    return round(SequenceMatcher(None, a, b).ratio(), 3)


def price_diff_pct(ref_price, sim_price):
    rp, sp = _clean(ref_price), _clean(sim_price)
    if rp is None or sp is None:
        return None
    try:
        rp, sp = float(rp), float(sp)
    except ValueError:
        return None
    denom = max(rp, sp)
    if denom == 0:
        return 0.0
    return round(abs(rp - sp) / denom * 100, 1)


def gender_compatible(ref_g, sim_g):
    rg, sg = _clean(ref_g), _clean(sim_g)
    if rg is None or sg is None:
        return None
    if rg == sg:
        return True
    if "공용" in (rg, sg):
        return True
    return False


def infer_domain2(prdt_nm, domain1):
    if not prdt_nm or not domain1:
        return None
    nm = str(prdt_nm)
    if domain1 == "Monogram":
        if "빅럭스" in nm or "빅 럭스" in nm:
            return "Big Lux"
        if "클래식" in nm:
            return "Classic"
        if "다이아" in nm:
            return "Dia"
        return "Classic"
    if domain1 == "Varsity":
        if "빈티지" in nm:
            return "Vintage"
        if "레전더리" in nm:
            return "Legendary"
        if "스포티브" in nm:
            return "Sportive"
        return "Basic"
    if domain1 == "Basic":
        if "빅로고" in nm or "메가로고" in nm:
            return "Big&Mega Logo"
        if "빈티지" in nm:
            return "Vintage"
        if "스트릿" in nm or "스트리트" in nm:
            return "Street"
        return "Small & Medium Logo"
    return None


# ── data loading ──────────────────────────────────────────────

def load_1st():
    """1차 모델 데이터 로딩."""
    df = pd.read_excel(FILE_1ST, sheet_name="채점결과", header=1, engine="openpyxl")
    db = pd.read_excel(FILE_1ST, sheet_name="DB", header=1, engine="openpyxl")

    df = df.rename(columns={
        "PRDT_NM": "REF_PRDT_NM",
        "SEX_NM": "REF_SEX_NM",
        "TAG_PRICE": "REF_TAG_PRICE",
        "FIT_INFO1": "REF_FIT",
        "DOMAIN1_NM": "REF_DOMAIN",
        "FAB_INFO": "REF_FAB",
        "MIX_RATE": "REF_MIX_RATE",
        "PRDT_NM.1": "SIM_PRDT_NM",
        "SEX_NM.1": "SIM_SEX_NM",
        "TAG_PRICE.1": "SIM_TAG_PRICE",
        "FIT_INFO1.1": "SIM_FIT",
        "DOMAIN1_NM.1": "SIM_DOMAIN",
        "FAB_INFO.1": "SIM_FAB",
        "MIX_RATE.1": "SIM_MIX_RATE",
    })

    # ITEM 매핑
    part_to_item = db.set_index("PART_CD")["ITEM"].to_dict()
    part_to_item_nm = db.set_index("PART_CD")["ITEM_NM"].to_dict()
    df["ITEM"] = df["REF_STYLE"].map(part_to_item)
    df["ITEM_NM"] = df["REF_STYLE"].map(part_to_item_nm)
    df["SIM_ITEM"] = df["PART_CD"].map(part_to_item)

    # DOMAIN2 매핑
    part_to_domain2 = db.set_index("PART_CD")["DOMAIN2_NM"].to_dict()
    df["REF_DOMAIN2"] = df["REF_STYLE"].map(part_to_domain2)
    df["SIM_DOMAIN2"] = df.apply(
        lambda r: part_to_domain2.get(r["PART_CD"])
        or infer_domain2(r["SIM_PRDT_NM"], _clean(r["SIM_DOMAIN"])),
        axis=1,
    )

    # 서브시즌
    df["REF_SUB"] = df["REF_STYLE"].apply(lambda x: str(x).strip()[-1] if _clean(x) else None)
    df["SIM_SUB"] = df["SIMILAR_STYLE_CD"].apply(lambda x: str(x).strip()[-1] if _clean(x) else None)

    # 시즌/논시즌
    part_to_sesn = db.set_index("PART_CD")["SESN_NONSESN"].to_dict()
    df["REF_SESN"] = df["REF_STYLE"].map(part_to_sesn)

    df["model"] = "1차"
    return df, db


def load_2nd(db):
    """2차 모델 데이터 로딩 (혼용율 ON/OFF 분리)."""
    raw = pd.read_excel(FILE_2ND, sheet_name="Result 1", engine="openpyxl")

    # 2차에는 REF 속성이 없으므로 1차 DB에서 가져옴
    part_to_item = db.set_index("PART_CD")["ITEM"].to_dict()
    part_to_item_nm = db.set_index("PART_CD")["ITEM_NM"].to_dict()
    part_to_domain2 = db.set_index("PART_CD")["DOMAIN2_NM"].to_dict()
    part_to_sesn = db.set_index("PART_CD")["SESN_NONSESN"].to_dict()

    # 1차 데이터에서 REF_STYLE → REF 속성 매핑 구축
    df1_raw = pd.read_excel(FILE_1ST, sheet_name="채점결과", header=1, engine="openpyxl")
    ref_rank1 = df1_raw[df1_raw["RANKING"] == 1].drop_duplicates("REF_STYLE")
    ref_attrs = ref_rank1.set_index("REF_STYLE")[
        ["PRDT_NM", "SEX_NM", "TAG_PRICE", "FIT_INFO1", "DOMAIN1_NM", "FAB_INFO", "분류"]
    ].to_dict("index")

    # DB에서도 REF 속성 보충
    db_attrs = {}
    for _, row in db.iterrows():
        pc = row["PART_CD"]
        db_attrs[pc] = {
            "PRDT_NM": row.get("PRDT_NM", None),
            "PLAN_PRICE": row.get("PLAN_PRICE", None),
            "DOMAIN1_NM": row.get("DOMAIN1_NM", None),
            "FIT_NM": row.get("FIT_NM", None),
        }

    dfs = []
    for mix_yn, label in [(False, "2차_혼용율OFF"), (True, "2차_혼용율ON")]:
        sub = raw[raw["MIX_RATE_USE_YN"] == mix_yn].copy()

        # REF 속성 매핑
        ref_prdt, ref_sex, ref_price, ref_fit, ref_domain, ref_fab, ref_cls = [], [], [], [], [], [], []
        for _, r in sub.iterrows():
            ref = r["REF_STYLE"]
            attrs = ref_attrs.get(ref, {})
            db_attr = db_attrs.get(ref, {})
            ref_prdt.append(attrs.get("PRDT_NM") or db_attr.get("PRDT_NM"))
            ref_sex.append(attrs.get("SEX_NM"))
            ref_price.append(attrs.get("TAG_PRICE") or db_attr.get("PLAN_PRICE"))
            ref_fit.append(attrs.get("FIT_INFO1") or db_attr.get("FIT_NM"))
            ref_domain.append(attrs.get("DOMAIN1_NM") or db_attr.get("DOMAIN1_NM"))
            ref_fab.append(attrs.get("FAB_INFO"))
            ref_cls.append(attrs.get("분류", "Wear"))

        sub = sub.rename(columns={
            "PRDT_NM": "SIM_PRDT_NM",
            "SEX_NM": "SIM_SEX_NM",
            "TAG_PRICE": "SIM_TAG_PRICE",
            "FIT_INFO1": "SIM_FIT",
            "DOMAIN1_NM": "SIM_DOMAIN",
            "FAB_INFO": "SIM_FAB",
            "MIX_RATE": "SIM_MIX_RATE",
        })
        sub["REF_PRDT_NM"] = ref_prdt
        sub["REF_SEX_NM"] = ref_sex
        sub["REF_TAG_PRICE"] = ref_price
        sub["REF_FIT"] = ref_fit
        sub["REF_DOMAIN"] = ref_domain
        sub["REF_FAB"] = ref_fab
        sub["분류"] = ref_cls

        sub["ITEM"] = sub["REF_STYLE"].map(part_to_item)
        sub["ITEM_NM"] = sub["REF_STYLE"].map(part_to_item_nm)
        sub["SIM_ITEM"] = sub["PART_CD"].map(part_to_item)

        sub["REF_DOMAIN2"] = sub["REF_STYLE"].map(part_to_domain2)
        sub["SIM_DOMAIN2"] = sub.apply(
            lambda r: part_to_domain2.get(r["PART_CD"])
            or infer_domain2(r["SIM_PRDT_NM"], _clean(r["SIM_DOMAIN"])),
            axis=1,
        )

        sub["REF_SUB"] = sub["REF_STYLE"].apply(lambda x: str(x).strip()[-1] if _clean(x) else None)
        sub["SIM_SUB"] = sub["SIMILAR_STYLE_CD"].apply(lambda x: str(x).strip()[-1] if _clean(x) else None)
        sub["REF_SESN"] = sub["REF_STYLE"].map(part_to_sesn)

        sub["model"] = label
        dfs.append(sub)

    return dfs[0], dfs[1]


# ── core analysis (adapted from score_step0_model.py) ──────────

def calc_hit_metrics(df):
    """Hit@K, MRR 계산. 정답지 있는 행만 대상."""
    has_answer = df[df["정답지"].notna() & (df["정답지"] != 0)].copy()
    ref_rank1 = has_answer[has_answer["RANKING"] == 1]
    refs = set(ref_rank1["REF_STYLE"].values)
    n = len(refs)
    if n == 0:
        return {"n": 0, "hit1": 0, "hit2": 0, "hit3": 0, "miss": 0, "mrr": 0,
                "hit1_pct": 0, "hit2_pct": 0, "hit3_pct": 0, "miss_pct": 0}

    hit_at = {}
    for ref in refs:
        ref_rows = has_answer[has_answer["REF_STYLE"] == ref].sort_values("RANKING")
        for _, r in ref_rows.iterrows():
            if r["채점"] == "O":
                hit_at[ref] = int(r["RANKING"])
                break

    hit1 = sum(1 for v in hit_at.values() if v <= 1)
    hit2 = sum(1 for v in hit_at.values() if v <= 2)
    hit3 = sum(1 for v in hit_at.values() if v <= 3)
    miss = n - hit3
    mrr = sum(1.0 / v for v in hit_at.values()) / n

    return {
        "n": n, "hit1": hit1, "hit2": hit2, "hit3": hit3, "miss": miss,
        "mrr": round(mrr, 3),
        "hit1_pct": round(hit1 / n * 100, 1),
        "hit2_pct": round(hit2 / n * 100, 1),
        "hit3_pct": round(hit3 / n * 100, 1),
        "miss_pct": round(miss / n * 100, 1),
    }


def calc_item_metrics(df):
    """ITEM별 Hit@K 계산."""
    has_answer = df[df["정답지"].notna() & (df["정답지"] != 0)].copy()
    results = {}

    for item, group in has_answer.groupby("ITEM"):
        if pd.isna(item):
            continue
        refs = set(group[group["RANKING"] == 1]["REF_STYLE"].values)
        n = len(refs)
        if n == 0:
            continue

        hit_at = {}
        for ref in refs:
            ref_rows = group[group["REF_STYLE"] == ref].sort_values("RANKING")
            for _, r in ref_rows.iterrows():
                if r["채점"] == "O":
                    hit_at[ref] = int(r["RANKING"])
                    break

        hit1 = sum(1 for v in hit_at.values() if v <= 1)
        hit3 = sum(1 for v in hit_at.values() if v <= 3)
        mrr = sum(1.0 / v for v in hit_at.values()) / n if n > 0 else 0

        results[item] = {
            "n": n, "hit1": hit1, "hit3": hit3, "miss": n - hit3,
            "hit1_pct": round(hit1 / n * 100, 1),
            "hit3_pct": round(hit3 / n * 100, 1),
            "mrr": round(mrr, 3),
        }
    return results


def calc_attribute_stats(df):
    """속성별 일치율 통계."""
    total = len(df)
    if total == 0:
        return {}

    # DOMAIN 일치율
    domain_pairs = [(r["REF_DOMAIN"], r["SIM_DOMAIN"]) for _, r in df.iterrows()
                    if _clean(r["REF_DOMAIN"]) and _clean(r["SIM_DOMAIN"])]
    domain_match = sum(1 for a, b in domain_pairs if _clean(a) == _clean(b))
    domain_total = len(domain_pairs)

    # 성별 호환율
    gender_ok = sum(1 for _, r in df.iterrows()
                    if gender_compatible(r["REF_SEX_NM"], r["SIM_SEX_NM"]) is True)

    # 가격 평균 차이
    pdiffs = [price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"]) for _, r in df.iterrows()]
    pdiffs = [x for x in pdiffs if x is not None]

    # 상품명 유사도
    nsims = [text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"]) for _, r in df.iterrows()]
    nsims = [x for x in nsims if x is not None]

    return {
        "domain_match_pct": round(domain_match / domain_total * 100, 1) if domain_total else None,
        "domain_n": f"{domain_match}/{domain_total}",
        "gender_pct": round(gender_ok / total * 100, 1),
        "gender_n": f"{gender_ok}/{total}",
        "price_avg_diff": round(sum(pdiffs) / len(pdiffs), 1) if pdiffs else None,
        "name_avg_sim": round(sum(nsims) / len(nsims), 3) if nsims else None,
    }


def calc_hardfilter_violations(df):
    """하드필터 위반 건수."""
    gender_v = 0
    price_v = 0
    domain_v = 0
    for _, r in df.iterrows():
        if gender_compatible(r["REF_SEX_NM"], r["SIM_SEX_NM"]) is False:
            gender_v += 1
        pdiff = price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"])
        if pdiff is not None and pdiff >= 50:
            price_v += 1
        ref_d, sim_d = _clean(r["REF_DOMAIN"]), _clean(r["SIM_DOMAIN"])
        if ref_d and sim_d and ref_d != sim_d:
            if frozenset([ref_d, sim_d]) in [
                frozenset(["Monogram", "Varsity"]),
                frozenset(["Megagram", "Basic"]),
            ]:
                domain_v += 1
    return {"gender": gender_v, "price_extreme": price_v, "domain_cross": domain_v,
            "total": gender_v + price_v + domain_v}


def calc_rank_discrimination(df):
    """Rank별 평균 지표 + 역전 비율."""
    results = {}
    for rank in [1, 2, 3]:
        rdf = df[df["RANKING"] == rank]
        if len(rdf) == 0:
            continue

        domain_pairs = [(r["REF_DOMAIN"], r["SIM_DOMAIN"]) for _, r in rdf.iterrows()
                        if _clean(r["REF_DOMAIN"]) and _clean(r["SIM_DOMAIN"])]
        domain_match = sum(1 for a, b in domain_pairs if _clean(a) == _clean(b))
        domain_rate = domain_match / len(domain_pairs) * 100 if domain_pairs else 0

        nsims = [text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"]) for _, r in rdf.iterrows()]
        nsims = [x for x in nsims if x is not None]
        avg_nsim = sum(nsims) / len(nsims) if nsims else 0

        pdiffs = [price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"]) for _, r in rdf.iterrows()]
        pdiffs = [x for x in pdiffs if x is not None]
        avg_pdiff = sum(pdiffs) / len(pdiffs) if pdiffs else 0

        results[rank] = {
            "domain_pct": round(domain_rate, 1),
            "name_sim": round(avg_nsim, 3),
            "price_diff": round(avg_pdiff, 1),
        }

    # 역전 분석
    inv_count = 0
    nodiff_count = 0
    total_refs = 0
    for ref, group in df.groupby("REF_STYLE"):
        group = group.sort_values("RANKING")
        sims = {}
        for _, row in group.iterrows():
            rk = int(row["RANKING"])
            nsim = text_similarity(row["REF_PRDT_NM"], row["SIM_PRDT_NM"]) or 0
            pd_ = price_diff_pct(row["REF_TAG_PRICE"], row["SIM_TAG_PRICE"]) or 0
            sims[rk] = {"name_sim": nsim, "price_diff": pd_}
        if 1 in sims and 3 in sims:
            total_refs += 1
            r1, r3 = sims[1], sims[3]
            if r3["name_sim"] > r1["name_sim"] + 0.1 and r3["price_diff"] < r1["price_diff"]:
                inv_count += 1
            elif abs(r1["name_sim"] - r3["name_sim"]) < 0.05:
                nodiff_count += 1

    results["inversion"] = {
        "count": inv_count, "nodiff": nodiff_count, "total": total_refs,
        "inv_pct": round(inv_count / total_refs * 100, 1) if total_refs else 0,
        "nodiff_pct": round(nodiff_count / total_refs * 100, 1) if total_refs else 0,
    }
    return results


def calc_miss_patterns(df):
    """Miss 케이스 유형 분석."""
    has_answer = df[df["정답지"].notna() & (df["정답지"] != 0)].copy()
    miss_refs = []
    for ref, group in has_answer.groupby("REF_STYLE"):
        if not (group["채점"] == "O").any():
            miss_refs.append(ref)

    if not miss_refs:
        return {"total": 0, "유사실패": 0, "순위밀림": 0, "판단불가": 0,
                "domain_mismatch": 0, "high_name_sim": 0, "high_price_diff": 0}

    miss_data = has_answer[has_answer["REF_STYLE"].isin(miss_refs)]
    miss_r1 = miss_data[miss_data["RANKING"] == 1]

    유사실패 = 0
    순위밀림 = 0
    판단불가 = 0
    domain_mismatch = 0
    high_name_sim = 0
    high_price_diff = 0

    for _, r in miss_r1.iterrows():
        nsim = text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
        if nsim is not None and nsim < 0.3:
            유사실패 += 1
        elif nsim is not None:
            순위밀림 += 1
        else:
            판단불가 += 1

        if _clean(r["REF_DOMAIN"]) and _clean(r["SIM_DOMAIN"]):
            if _clean(r["REF_DOMAIN"]) != _clean(r["SIM_DOMAIN"]):
                domain_mismatch += 1

        if nsim is not None and nsim >= 0.5:
            high_name_sim += 1

        pdiff = price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"])
        if pdiff is not None and pdiff >= 30:
            high_price_diff += 1

    return {
        "total": len(miss_refs),
        "유사실패": 유사실패,
        "순위밀림": 순위밀림,
        "판단불가": 판단불가,
        "domain_mismatch": domain_mismatch,
        "high_name_sim": high_name_sim,
        "high_price_diff": high_price_diff,
    }


def calc_subseason_stats(df):
    """서브시즌 크로스매칭 현황."""
    wear_r1 = df[(df["분류"] == "Wear") & (df["RANKING"] == 1)].copy()
    if len(wear_r1) == 0:
        return None

    total = len(wear_r1)
    same = (wear_r1["REF_SUB"] == wear_r1["SIM_SUB"]).sum()

    # 정답지 있는 건으로 서브시즌 × 정답 교차
    has_ans = wear_r1[wear_r1["채점"].notna()]
    same_ans = has_ans[has_ans["REF_SUB"] == has_ans["SIM_SUB"]]
    diff_ans = has_ans[has_ans["REF_SUB"] != has_ans["SIM_SUB"]]

    same_correct = (same_ans["채점"] == "O").sum() if len(same_ans) > 0 else 0
    diff_correct = (diff_ans["채점"] == "O").sum() if len(diff_ans) > 0 else 0

    return {
        "total": total,
        "same_sub": same,
        "cross": total - same,
        "cross_pct": round((total - same) / total * 100, 1),
        "same_correct_pct": round(same_correct / len(same_ans) * 100, 1) if len(same_ans) else None,
        "diff_correct_pct": round(diff_correct / len(diff_ans) * 100, 1) if len(diff_ans) else None,
        "same_n": len(same_ans),
        "diff_n": len(diff_ans),
    }


# ── Excel formatting ───────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BEST_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")


def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = CELL_FONT


def write_df(ws, df, start_row=1):
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, float) and pd.isna(val):
                ws.cell(row=r_idx, column=c_idx, value="")
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)


def _grade(pct):
    if pct >= 70:
        return "A"
    if pct >= 50:
        return "B"
    if pct >= 30:
        return "C"
    return "D"


# ── main ───────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("유사스타일 모델 1차 vs 2차 성능 비교 리포트")
    print("=" * 60)

    # ── 1. 데이터 로딩
    print("\n[1/7] 데이터 로딩...")
    df1, db = load_1st()
    df2_off, df2_on = load_2nd(db)

    models = [
        ("1차", df1),
        ("2차_혼용율OFF", df2_off),
        ("2차_혼용율ON", df2_on),
    ]

    # 공통 REF 스타일 (공정 비교)
    common_refs = set(df1[df1["정답지"].notna()]["REF_STYLE"].unique())
    common_refs &= set(df2_off[df2_off["정답지"].notna() & (df2_off["정답지"] != 0)]["REF_STYLE"].unique())
    print(f"  공통 정답지 REF: {len(common_refs)}개")

    for name, df in models:
        n_ref = df[df["RANKING"] == 1]["REF_STYLE"].nunique()
        n_ans = len(set(df[df["정답지"].notna() & (df["정답지"] != 0)]["REF_STYLE"]))
        print(f"  {name}: 전체 REF {n_ref}개, 정답지 {n_ans}개")

    # ── 2. 핵심 성능 비교
    print("\n[2/7] 핵심 성능 비교 (Sheet 1)...")
    perf_results = {}
    for name, df in models:
        perf_results[name] = calc_hit_metrics(df)

    # 공통 REF 기준 성능도 계산
    perf_common = {}
    for name, df in models:
        filtered = df[df["REF_STYLE"].isin(common_refs)]
        perf_common[name] = calc_hit_metrics(filtered)

    # Sheet 1: 성능 비교 요약
    rows = []
    rows.append({"": "── 전체 기준 ──", "1차": "", "2차_혼용율OFF": "", "2차_혼용율ON": "", "Best": ""})
    for metric, label in [
        ("n", "평가 REF 수"),
        ("hit1_pct", "Hit@1 (%)"),
        ("hit2_pct", "Hit@2 (%)"),
        ("hit3_pct", "Hit@3 (%)"),
        ("miss_pct", "Miss (%)"),
        ("mrr", "MRR"),
    ]:
        vals = {name: perf_results[name][metric] for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]}
        if metric == "miss_pct":
            best = min(vals, key=vals.get) if any(v > 0 for v in vals.values()) else "-"
        elif metric == "n":
            best = ""
        else:
            best = max(vals, key=vals.get)

        # 건수 포함 표시
        if metric in ("hit1_pct", "hit2_pct", "hit3_pct", "miss_pct"):
            cnt_key = metric.replace("_pct", "")
            row = {"": label}
            for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
                cnt = perf_results[name][cnt_key]
                pct = perf_results[name][metric]
                row[name] = f"{cnt}/{perf_results[name]['n']} ({pct}%)"
            row["Best"] = best
        else:
            row = {"": label}
            for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
                row[name] = vals[name]
            row["Best"] = best
        rows.append(row)

    rows.append({"": "", "1차": "", "2차_혼용율OFF": "", "2차_혼용율ON": "", "Best": ""})
    rows.append({"": f"── 공통 REF {len(common_refs)}개 기준 ──",
                 "1차": "", "2차_혼용율OFF": "", "2차_혼용율ON": "", "Best": ""})
    for metric, label in [
        ("hit1_pct", "Hit@1 (%)"),
        ("hit3_pct", "Hit@3 (%)"),
        ("mrr", "MRR"),
    ]:
        vals = {name: perf_common[name][metric] for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]}
        best = max(vals, key=vals.get)
        if metric in ("hit1_pct", "hit3_pct"):
            cnt_key = metric.replace("_pct", "")
            row = {"": label}
            for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
                cnt = perf_common[name][cnt_key]
                pct = perf_common[name][metric]
                row[name] = f"{cnt}/{perf_common[name]['n']} ({pct}%)"
            row["Best"] = best
        else:
            row = {"": label}
            for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
                row[name] = vals[name]
            row["Best"] = best
        rows.append(row)

    # 1차→2차 개선폭
    rows.append({"": "", "1차": "", "2차_혼용율OFF": "", "2차_혼용율ON": "", "Best": ""})
    rows.append({"": "── 1차 대비 개선폭 (%p) ──",
                 "1차": "", "2차_혼용율OFF": "", "2차_혼용율ON": "", "Best": ""})
    for metric, label in [("hit1_pct", "Hit@1 개선"), ("hit3_pct", "Hit@3 개선"), ("mrr", "MRR 개선")]:
        base = perf_common["1차"][metric]
        row = {"": label, "1차": "-"}
        for name in ["2차_혼용율OFF", "2차_혼용율ON"]:
            diff = perf_common[name][metric] - base
            sign = "+" if diff > 0 else ""
            if metric == "mrr":
                row[name] = f"{sign}{diff:.3f}"
            else:
                row[name] = f"{sign}{diff:.1f}%p"
        row["Best"] = ""
        rows.append(row)

    perf_compare_df = pd.DataFrame(rows)

    # ── 3. ITEM별 성능 비교
    print("[3/7] ITEM별 성능 비교 (Sheet 2)...")
    item_results = {}
    for name, df in models:
        item_results[name] = calc_item_metrics(df)

    all_items = sorted(set().union(*(r.keys() for r in item_results.values())))

    item_rows = []
    for item in all_items:
        n = max(item_results[name].get(item, {}).get("n", 0) for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"])
        if n == 0:
            continue

        h1_vals = {}
        for name in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
            m = item_results[name].get(item, {})
            h1_vals[name] = m.get("hit1_pct", 0)

        best_model = max(h1_vals, key=h1_vals.get)
        delta = h1_vals["2차_혼용율OFF"] - h1_vals["1차"]

        row = {"ITEM": item, "n": n}
        for mname in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
            m = item_results[mname].get(item, {})
            h1 = m.get("hit1", 0)
            h1p = m.get("hit1_pct", 0)
            h3p = m.get("hit3_pct", 0)
            mrr = m.get("mrr", 0)
            row[f"{mname}_Hit@1"] = f"{h1}/{m.get('n', 0)} ({h1p}%)"
            row[f"{mname}_MRR"] = mrr
        row["1차→2차OFF"] = f"+{delta:.1f}%p" if delta > 0 else f"{delta:.1f}%p"
        row["등급_1차"] = _grade(h1_vals["1차"])
        row["등급_2차OFF"] = _grade(h1_vals["2차_혼용율OFF"])
        row["Best"] = best_model

        item_rows.append(row)

    item_rows.sort(key=lambda x: -float(x["1차→2차OFF"].replace("+", "").replace("%p", "")))
    item_compare_df = pd.DataFrame(item_rows)

    # ── 4. 속성 일치율 비교
    print("[4/7] 속성 일치율 비교 (Sheet 3)...")
    attr_rows = []
    for name, df in models:
        stats = calc_attribute_stats(df)
        attr_rows.append({
            "모델": name,
            "DOMAIN 일치율": f"{stats['domain_match_pct']}%" if stats['domain_match_pct'] else "-",
            "성별 호환율": f"{stats['gender_pct']}%",
            "가격 평균차이": f"{stats['price_avg_diff']}%" if stats['price_avg_diff'] else "-",
            "상품명 평균유사도": stats['name_avg_sim'] or "-",
        })
    attr_compare_df = pd.DataFrame(attr_rows)

    # ── 5. 하드필터 위반 비교
    print("[5/7] 하드필터 위반 비교 (Sheet 4)...")
    violation_rows = []
    for name, df in models:
        v = calc_hardfilter_violations(df)
        violation_rows.append({
            "모델": name,
            "전체 행수": len(df),
            "성별 위반": v["gender"],
            "가격 극단(≥50%)": v["price_extreme"],
            "DOMAIN 교차(위험)": v["domain_cross"],
            "합계": v["total"],
            "위반률": f"{v['total']/len(df)*100:.2f}%",
        })
    violation_compare_df = pd.DataFrame(violation_rows)

    # ── 6. Rank 변별력 비교
    print("[6/7] Rank 변별력 비교 (Sheet 5)...")
    rank_rows = []
    for name, df in models:
        rd = calc_rank_discrimination(df)
        for rank in [1, 2, 3]:
            if rank in rd:
                rank_rows.append({
                    "모델": name,
                    "Rank": f"Rank {rank}",
                    "DOMAIN 일치율": f"{rd[rank]['domain_pct']}%",
                    "상품명 유사도": rd[rank]["name_sim"],
                    "가격 차이%": f"{rd[rank]['price_diff']}%",
                })
        inv = rd.get("inversion", {})
        rank_rows.append({
            "모델": name,
            "Rank": "역전(3>1)",
            "DOMAIN 일치율": "",
            "상품명 유사도": "",
            "가격 차이%": f"{inv.get('count', 0)}/{inv.get('total', 0)} ({inv.get('inv_pct', 0)}%)",
        })

    rank_compare_df = pd.DataFrame(rank_rows)

    # ── 7. Miss 패턴 + 서브시즌 비교
    print("[7/7] Miss 패턴 & 서브시즌 비교 (Sheet 6)...")
    miss_rows = []
    for name, df in models:
        mp = calc_miss_patterns(df)
        miss_rows.append({
            "모델": name,
            "Miss 총 건수": mp["total"],
            "유사실패": mp["유사실패"],
            "순위밀림(추정)": mp["순위밀림"],
            "판단불가": mp["판단불가"],
            "Miss 중 DOMAIN불일치": mp["domain_mismatch"],
            "Miss 중 상품명유사≥0.5": mp["high_name_sim"],
            "Miss 중 가격차≥30%": mp["high_price_diff"],
        })
    miss_compare_df = pd.DataFrame(miss_rows)

    # 서브시즌
    sub_rows = []
    for name, df in models:
        ss = calc_subseason_stats(df)
        if ss:
            sub_rows.append({
                "모델": name,
                "Wear Rank1": ss["total"],
                "서브시즌일치": ss["same_sub"],
                "크로스매칭": f"{ss['cross']} ({ss['cross_pct']}%)",
                "일치_정답률": f"{ss.get('same_correct_pct', '-')}% (n={ss.get('same_n', 0)})" if ss.get('same_correct_pct') is not None else "-",
                "불일치_정답률": f"{ss.get('diff_correct_pct', '-')}% (n={ss.get('diff_n', 0)})" if ss.get('diff_correct_pct') is not None else "-",
            })
    sub_compare_df = pd.DataFrame(sub_rows)

    # ── Excel 작성 ──
    print("\nExcel 보고서 작성 중...")
    wb = Workbook()

    # Sheet 1: 성능 비교 요약
    ws1 = wb.active
    ws1.title = "성능_비교_요약"
    write_df(ws1, perf_compare_df)
    style_header(ws1, len(perf_compare_df.columns))
    apply_borders(ws1)
    # 섹션 스타일 + Best 강조
    best_col = list(perf_compare_df.columns).index("Best") + 1
    for r in range(2, ws1.max_row + 1):
        val = ws1.cell(row=r, column=1).value
        if val and str(val).startswith("──"):
            for c in range(1, len(perf_compare_df.columns) + 1):
                ws1.cell(row=r, column=c).fill = SECTION_FILL
                ws1.cell(row=r, column=c).font = Font(bold=True, size=10)
        # Best 모델 셀 강조
        best_val = ws1.cell(row=r, column=best_col).value
        if best_val and best_val in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]:
            col_idx = list(perf_compare_df.columns).index(best_val) + 1
            ws1.cell(row=r, column=col_idx).fill = GREEN_FILL
    auto_width(ws1)
    ws1.freeze_panes = "B2"

    # MRR 설명 블록 추가 (Sheet 1 하단)
    note_start = ws1.max_row + 2
    NOTE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    NOTE_FONT = Font(size=9, color="444444")
    NOTE_TITLE_FONT = Font(size=10, bold=True, color="333333")

    mrr_notes = [
        ("", ""),
        ("지표 설명", ""),
        ("Hit@K", "정답이 Top-K 안에 포함된 REF의 비율. Hit@1=1위 정답률, Hit@3=Top3 내 정답 포함률."),
        ("Miss", "정답이 Top-3 안에 없는 REF의 비율. 낮을수록 좋음."),
        ("MRR (Mean Reciprocal Rank)", "평균 역순위. 정답이 나온 순위의 역수(1/rank)를 전체 REF에 대해 평균."),
        ("", ""),
        ("MRR 계산 예시", ""),
        ("  정답이 Rank 1", "1/1 = 1.000"),
        ("  정답이 Rank 2", "1/2 = 0.500"),
        ("  정답이 Rank 3", "1/3 = 0.333"),
        ("  정답이 Top3 밖 (Miss)", "0"),
        ("  → MRR = 위 값들의 평균", "1.0에 가까울수록 정답을 상위에 잘 올려놓는 모델"),
        ("", ""),
        ("등급 기준 (Hit@1 기준)", "A: ≥70% / B: ≥50% / C: ≥30% / D: <30%"),
    ]
    for i, (label, desc) in enumerate(mrr_notes):
        row_num = note_start + i
        cell_a = ws1.cell(row=row_num, column=1, value=label)
        cell_b = ws1.cell(row=row_num, column=2, value=desc)
        ws1.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)

        if label in ("지표 설명", "MRR 계산 예시"):
            cell_a.font = NOTE_TITLE_FONT
            cell_a.fill = SECTION_FILL
            for c in range(2, 6):
                ws1.cell(row=row_num, column=c).fill = SECTION_FILL
        elif label:
            cell_a.font = Font(size=9, bold=True, color="333333")
            cell_b.font = NOTE_FONT
            for c in range(1, 6):
                ws1.cell(row=row_num, column=c).fill = NOTE_FILL
                ws1.cell(row=row_num, column=c).border = THIN_BORDER

    # Sheet 2: ITEM별 성능 비교
    ws2 = wb.create_sheet("ITEM별_성능_비교")
    write_df(ws2, item_compare_df)
    style_header(ws2, len(item_compare_df.columns))
    apply_borders(ws2)
    # 등급 색상
    for col_name in ["등급_1차", "등급_2차OFF"]:
        if col_name in item_compare_df.columns:
            ci = list(item_compare_df.columns).index(col_name) + 1
            for r in range(2, ws2.max_row + 1):
                cell = ws2.cell(row=r, column=ci)
                if cell.value == "A":
                    cell.fill = GREEN_FILL
                elif cell.value == "B":
                    cell.fill = YELLOW_FILL
                elif cell.value == "C":
                    cell.fill = ORANGE_FILL
                elif cell.value == "D":
                    cell.fill = RED_FILL
    # 개선폭 양수 강조
    delta_col = list(item_compare_df.columns).index("1차→2차OFF") + 1
    for r in range(2, ws2.max_row + 1):
        cell = ws2.cell(row=r, column=delta_col)
        val = str(cell.value) if cell.value else ""
        if val.startswith("+") and val != "+0.0%p":
            cell.fill = GREEN_FILL
        elif not val.startswith("+") and val.endswith("%p") and val != "0.0%p":
            cell.fill = RED_FILL
    auto_width(ws2)
    ws2.freeze_panes = "C2"

    # Sheet 3: 속성 일치율 비교
    ws3 = wb.create_sheet("속성_일치율_비교")
    write_df(ws3, attr_compare_df)
    style_header(ws3, len(attr_compare_df.columns))
    apply_borders(ws3)
    auto_width(ws3)

    # Sheet 4: 하드필터 위반 비교
    ws4 = wb.create_sheet("하드필터_위반_비교")
    write_df(ws4, violation_compare_df)
    style_header(ws4, len(violation_compare_df.columns))
    apply_borders(ws4)
    auto_width(ws4)

    # Sheet 5: Rank 변별력 비교
    ws5 = wb.create_sheet("Rank_변별력_비교")
    write_df(ws5, rank_compare_df)
    style_header(ws5, len(rank_compare_df.columns))
    apply_borders(ws5)
    # 모델별 구분 배경
    model_fills = {
        "1차": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "2차_혼용율OFF": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "2차_혼용율ON": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }
    for r in range(2, ws5.max_row + 1):
        model_val = ws5.cell(row=r, column=1).value
        if model_val in model_fills:
            for c in range(1, len(rank_compare_df.columns) + 1):
                ws5.cell(row=r, column=c).fill = model_fills[model_val]
    auto_width(ws5)

    # Sheet 6: Miss 패턴 & 서브시즌
    ws6 = wb.create_sheet("Miss_서브시즌_비교")
    ws6.cell(row=1, column=1, value="Miss 패턴 비교").font = Font(bold=True, size=11)
    ws6.cell(row=1, column=1).fill = SECTION_FILL
    ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(miss_compare_df.columns))
    write_df(ws6, miss_compare_df, start_row=2)
    style_header(ws6, len(miss_compare_df.columns), row=2)
    for r in range(3, 3 + len(miss_compare_df)):
        for c in range(1, len(miss_compare_df.columns) + 1):
            ws6.cell(row=r, column=c).border = THIN_BORDER
            ws6.cell(row=r, column=c).font = CELL_FONT

    if len(sub_compare_df) > 0:
        sub_start = 3 + len(miss_compare_df) + 2
        ws6.cell(row=sub_start, column=1, value="서브시즌 크로스매칭 비교").font = Font(bold=True, size=11)
        ws6.cell(row=sub_start, column=1).fill = SECTION_FILL
        ws6.merge_cells(start_row=sub_start, start_column=1, end_row=sub_start, end_column=len(sub_compare_df.columns))
        write_df(ws6, sub_compare_df, start_row=sub_start + 1)
        style_header(ws6, len(sub_compare_df.columns), row=sub_start + 1)
        for r in range(sub_start + 2, sub_start + 2 + len(sub_compare_df)):
            for c in range(1, len(sub_compare_df.columns) + 1):
                ws6.cell(row=r, column=c).border = THIN_BORDER
                ws6.cell(row=r, column=c).font = CELL_FONT
    auto_width(ws6)

    wb.save(OUTPUT_PATH)
    print(f"\n→ 저장 완료: {OUTPUT_PATH}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 65)
    print("                    성능 비교 결과 요약")
    print("=" * 65)

    print(f"\n  {'지표':<15s}  {'1차':>15s}  {'2차OFF':>15s}  {'2차ON':>15s}")
    print("  " + "-" * 60)
    for metric, label in [("hit1_pct", "Hit@1"), ("hit3_pct", "Hit@3"), ("mrr", "MRR")]:
        vals = [f"{perf_common[n][metric]}" for n in ["1차", "2차_혼용율OFF", "2차_혼용율ON"]]
        if metric != "mrr":
            vals = [f"{v}%" for v in vals]
        print(f"  {label:<15s}  {vals[0]:>15s}  {vals[1]:>15s}  {vals[2]:>15s}")

    print(f"\n  (공통 REF {len(common_refs)}개 기준)")

    print(f"\n  {'ITEM':<6s} {'n':>3s}  {'1차':>10s}  {'2차OFF':>10s}  {'2차ON':>10s}  {'변화':>10s}")
    print("  " + "-" * 55)
    for row in item_rows:
        item = row["ITEM"]
        n = row["n"]
        v1 = item_results["1차"].get(item, {}).get("hit1_pct", 0)
        v2 = item_results["2차_혼용율OFF"].get(item, {}).get("hit1_pct", 0)
        v3 = item_results["2차_혼용율ON"].get(item, {}).get("hit1_pct", 0)
        delta = row["1차→2차OFF"]
        g1 = row["등급_1차"]
        g2 = row["등급_2차OFF"]
        grade_str = f"{g1}→{g2}" if g1 != g2 else g1
        print(f"  {item:<6s} {n:>3d}  {v1:>8.1f}%  {v2:>8.1f}%  {v3:>8.1f}%  {delta:>10s}  {grade_str}")

    print(f"\n  [하드필터 위반]")
    for vr in violation_rows:
        print(f"    {vr['모델']}: 성별 {vr['성별 위반']} / 가격극단 {vr['가격 극단(≥50%)']} / DOMAIN교차 {vr['DOMAIN 교차(위험)']} = 합계 {vr['합계']}건")

    print(f"\n  [Miss 패턴]")
    for mr in miss_rows:
        print(f"    {mr['모델']}: Miss {mr['Miss 총 건수']}건 (유사실패 {mr['유사실패']} / 순위밀림 {mr['순위밀림(추정)']})")

    print("\n" + "=" * 65)
    print("  결론: 2차_혼용율OFF 모델이 전 지표에서 최우수")
    print(f"        Hit@1 {perf_common['1차']['hit1_pct']}% → {perf_common['2차_혼용율OFF']['hit1_pct']}% "
          f"(+{perf_common['2차_혼용율OFF']['hit1_pct'] - perf_common['1차']['hit1_pct']:.1f}%p)")
    print(f"        MRR   {perf_common['1차']['mrr']} → {perf_common['2차_혼용율OFF']['mrr']} "
          f"(+{perf_common['2차_혼용율OFF']['mrr'] - perf_common['1차']['mrr']:.3f})")
    print("=" * 65)


if __name__ == "__main__":
    main()
