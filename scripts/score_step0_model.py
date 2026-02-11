"""
STEP 0 유사스타일 모델 성능 측정 리포트

정답지(사업부 기획 시 사용하는 유사스타일 매핑)와 모델 반환값을 대조하여
정량적 성능 측정을 수행하고, 정답지 없는 스타일의 성능도 간접 추정한다.

- Input:  output/유사스타일자동반환결과+정답지.xlsx (채점결과, 정답지, DB 시트)
- Output: output/step0_model_performance.xlsx (12 sheets) + 콘솔 요약
"""

import os
import re
from collections import defaultdict
from difflib import SequenceMatcher

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_PATH = os.path.join(BASE_DIR, "output", "유사스타일자동반환결과+정답지.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "step0_model_performance.xlsx")


# ── helpers ────────────────────────────────────────────────────

def _clean(v):
    """None / 0 / 'None' / NaN → None, else stripped string."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    if s in ("", "0", "None", "nan", "NaN"):
        return None
    return s


def text_similarity(a, b):
    a, b = _clean(a), _clean(b)
    if a is None or b is None:
        return None
    a = re.sub(r"[^\w\s]", "", a)
    b = re.sub(r"[^\w\s]", "", b)
    return round(SequenceMatcher(None, a, b).ratio(), 3)


def price_diff_pct(ref_price, sim_price):
    rp, sp = _clean(ref_price), _clean(sim_price)
    if rp is None or sp is None:
        return None
    try:
        rp, sp = float(rp), float(sp)
    except ValueError:
        return None
    denom = max(rp, sp)
    if denom == 0:
        return 0.0
    return round(abs(rp - sp) / denom * 100, 1)


def gender_compatible(ref_g, sim_g):
    rg, sg = _clean(ref_g), _clean(sim_g)
    if rg is None or sg is None:
        return None
    if rg == sg:
        return True
    if "공용" in (rg, sg):
        return True
    return False


def infer_domain2(prdt_nm, domain1):
    """상품명 키워드 + DOMAIN1 기반으로 DOMAIN2 추정."""
    if not prdt_nm or not domain1:
        return None
    nm = str(prdt_nm)

    if domain1 == "Monogram":
        if "빅럭스" in nm or "빅 럭스" in nm:
            return "Big Lux"
        if "클래식" in nm:
            return "Classic"
        if "다이아" in nm:
            return "Dia"
        if "럭셔리" in nm:
            return "Luxury"
        if "모노티브" in nm:
            return "Monotive"
        return "Classic"  # Monogram 기본값

    if domain1 == "Varsity":
        if "빈티지" in nm:
            return "Vintage"
        if "레전더리" in nm:
            return "Legendary"
        if "스포티브" in nm:
            return "Sportive"
        if "베이직" in nm:
            return "Basic"
        return "Basic"  # Varsity 기본값

    if domain1 == "Basic":
        if "힙" in nm and "아웃도어" in nm:
            return "Hip Outdoor"
        if "빅로고" in nm or "메가로고" in nm or "빅 로고" in nm:
            return "Big&Mega Logo"
        if "빈티지" in nm:
            return "Vintage"
        if "스트릿" in nm or "스트리트" in nm:
            return "Street"
        return "Small & Medium Logo"  # Basic 기본값

    if domain1 == "Heart":
        return "Street"

    return None


def infer_sesn(style_cd):
    """스타일코드 끝자리로 시즌/논시즌 추정."""
    if not style_cd:
        return None
    last = str(style_cd).strip()[-1].upper()
    return "논시즌" if last == "N" else "시즌"


# ── data loading ───────────────────────────────────────────────

def load_data():
    """채점결과 + DB 시트를 로드하고 ITEM/서브시즌/DOMAIN2/시즌 매핑."""
    df = pd.read_excel(INPUT_PATH, sheet_name="채점결과", header=1, engine="openpyxl")
    db = pd.read_excel(INPUT_PATH, sheet_name="DB", header=1, engine="openpyxl")

    # REF 쪽 컬럼과 SIM 쪽 컬럼 구분을 위해 rename
    df = df.rename(columns={
        "PRDT_NM": "REF_PRDT_NM",
        "SEX_NM": "REF_SEX_NM",
        "TAG_PRICE": "REF_TAG_PRICE",
        "FIT_INFO1": "REF_FIT",
        "DOMAIN1_NM": "REF_DOMAIN",
        "FAB_INFO": "REF_FAB",
        "MIX_RATE": "REF_MIX_RATE",
        "PRDT_NM.1": "SIM_PRDT_NM",
        "SEX_NM.1": "SIM_SEX_NM",
        "TAG_PRICE.1": "SIM_TAG_PRICE",
        "FIT_INFO1.1": "SIM_FIT",
        "DOMAIN1_NM.1": "SIM_DOMAIN",
        "FAB_INFO.1": "SIM_FAB",
        "MIX_RATE.1": "SIM_MIX_RATE",
    })

    # PART_CD → ITEM 매핑 (DB 시트)
    part_to_item = db.set_index("PART_CD")["ITEM"].to_dict()
    part_to_item_nm = db.set_index("PART_CD")["ITEM_NM"].to_dict()
    df["ITEM"] = df["REF_STYLE"].map(part_to_item)
    df["ITEM_NM"] = df["REF_STYLE"].map(part_to_item_nm)

    # SIM 쪽도 ITEM 매핑
    df["SIM_ITEM"] = df["PART_CD"].map(part_to_item)

    # ── 서브시즌 (스타일코드 끝1자리) ──
    df["REF_SUB"] = df["REF_STYLE"].apply(lambda x: str(x).strip()[-1] if _clean(x) else None)
    df["SIM_SUB"] = df["SIMILAR_STYLE_CD"].apply(
        lambda x: str(x).strip()[-1] if _clean(x) else None
    )

    # ── DOMAIN2 매핑 ──
    part_to_domain2 = db.set_index("PART_CD")["DOMAIN2_NM"].to_dict()
    df["REF_DOMAIN2"] = df["REF_STYLE"].map(part_to_domain2)
    # SIM DOMAIN2: DB에 없으므로 상품명+DOMAIN1 기반 추정
    df["SIM_DOMAIN2"] = df.apply(
        lambda r: part_to_domain2.get(r["PART_CD"])
        or infer_domain2(r["SIM_PRDT_NM"], _clean(r["SIM_DOMAIN"])),
        axis=1,
    )

    # ── 시즌/논시즌 매핑 ──
    part_to_sesn = db.set_index("PART_CD")["SESN_NONSESN"].to_dict()
    df["REF_SESN"] = df["REF_STYLE"].map(part_to_sesn)
    # SIM 시즌: PART_CD 기반 or 스타일코드 끝자리 추정
    df["SIM_SESN"] = df["PART_CD"].map(part_to_sesn).fillna(
        df["PART_CD"].apply(infer_sesn)
    )

    return df, db


# ── Sheet 1: 성능_요약 ────────────────────────────────────────

def build_performance_summary(df):
    """정답지 기반 핵심 성능 지표."""
    # 정답지 있는 REF만 (Rank 1 기준으로 REF 목록)
    has_answer = df[df["정답지"].notna()].copy()
    ref_rank1 = has_answer[has_answer["RANKING"] == 1]
    total_refs = len(ref_rank1)

    def calc_metrics(subset_df, all_ranks_df):
        """subset_df = Rank1 rows, all_ranks_df = 해당 REF의 전체 Rank rows."""
        n = len(subset_df)
        if n == 0:
            return {"Total": 0, "Hit@1": "-", "Hit@2": "-", "Hit@3": "-",
                    "Miss": "-", "MRR": "-"}

        ref_styles = set(subset_df["REF_STYLE"].values)
        filtered = all_ranks_df[all_ranks_df["REF_STYLE"].isin(ref_styles)]

        hit1 = len(filtered[(filtered["RANKING"] == 1) & (filtered["채점"] == "O")])
        hit2_extra = len(filtered[(filtered["RANKING"] == 2) & (filtered["채점"] == "O")])
        hit3_extra = len(filtered[(filtered["RANKING"] == 3) & (filtered["채점"] == "O")])

        # Hit@K = REF 중 top-K 안에 정답이 있는 비율
        hit_at = {}
        for ref in ref_styles:
            ref_rows = filtered[filtered["REF_STYLE"] == ref].sort_values("RANKING")
            for k in [1, 2, 3]:
                top_k = ref_rows[ref_rows["RANKING"] <= k]
                if (top_k["채점"] == "O").any():
                    hit_at.setdefault(ref, k)

        hit1_count = sum(1 for v in hit_at.values() if v <= 1)
        hit2_count = sum(1 for v in hit_at.values() if v <= 2)
        hit3_count = sum(1 for v in hit_at.values() if v <= 3)
        miss_count = n - hit3_count

        # MRR
        mrr_sum = 0
        for ref in ref_styles:
            if ref in hit_at:
                mrr_sum += 1.0 / hit_at[ref]
        mrr = mrr_sum / n

        return {
            "Total": n,
            "Hit@1": f"{hit1_count}/{n} ({hit1_count/n*100:.1f}%)",
            "Hit@2": f"{hit2_count}/{n} ({hit2_count/n*100:.1f}%)",
            "Hit@3": f"{hit3_count}/{n} ({hit3_count/n*100:.1f}%)",
            "Miss": f"{miss_count}/{n} ({miss_count/n*100:.1f}%)",
            "MRR": f"{mrr:.3f}",
        }

    # 전체 성능
    overall = calc_metrics(ref_rank1, has_answer)
    overall["구분"] = "전체"

    # ITEM별 성능
    item_rows = []
    for item, group in ref_rank1.groupby("ITEM"):
        if pd.isna(item):
            continue
        item_nm = group["ITEM_NM"].iloc[0] if "ITEM_NM" in group.columns else ""
        all_ranks = has_answer[has_answer["REF_STYLE"].isin(group["REF_STYLE"])]
        m = calc_metrics(group, all_ranks)
        m["구분"] = f"{item}({item_nm})" if _clean(item_nm) else str(item)
        item_rows.append(m)

    # Hit@1 기준 내림차순 정렬 (Total 숫자로 파싱)
    def sort_key(row):
        h1 = row["Hit@1"]
        if h1 == "-":
            return -1
        pct = float(h1.split("(")[1].replace("%)", ""))
        return pct
    item_rows.sort(key=sort_key, reverse=True)

    # 성능 등급 부여
    for row in item_rows:
        h1 = row["Hit@1"]
        if h1 == "-" or row["Total"] == 0:
            row["등급"] = "-"
            continue
        pct = float(h1.split("(")[1].replace("%)", ""))
        if pct >= 70:
            row["등급"] = "A"
        elif pct >= 50:
            row["등급"] = "B"
        elif pct >= 30:
            row["등급"] = "C"
        else:
            row["등급"] = "D"

    overall["등급"] = ""
    all_rows = [overall] + item_rows

    cols = ["구분", "Total", "Hit@1", "Hit@2", "Hit@3", "Miss", "MRR", "등급"]
    return pd.DataFrame(all_rows)[cols]


# ── Sheet 2: Miss_케이스_상세 ──────────────────────────────────

def build_miss_analysis(df):
    """정답이 Top 3에 없는 케이스 전수 분석."""
    has_answer = df[df["정답지"].notna()].copy()

    # REF별로 채점 결과 확인 - Hit@3에 해당하지 않는 REF = Miss
    miss_refs = []
    for ref, group in has_answer.groupby("REF_STYLE"):
        if not (group["채점"] == "O").any():
            miss_refs.append(ref)

    if not miss_refs:
        return pd.DataFrame()

    rows = []
    for ref in miss_refs:
        ref_rows = has_answer[has_answer["REF_STYLE"] == ref].sort_values("RANKING")
        ref_info = ref_rows.iloc[0]
        answer = ref_info["정답지"]

        # Rank 1~3 반환값
        rank_styles = {}
        rank_names = {}
        rank_domains = {}
        rank_fits = {}
        rank_prices = {}
        for _, r in ref_rows.iterrows():
            rk = int(r["RANKING"])
            rank_styles[rk] = r["SIMILAR_STYLE_CD"]
            rank_names[rk] = r["SIM_PRDT_NM"]
            rank_domains[rk] = _clean(r["SIM_DOMAIN"])
            rank_fits[rk] = _clean(r["SIM_FIT"])
            rank_prices[rk] = r["SIM_TAG_PRICE"]

        # 정답 vs Rank1 비교
        r1_name_sim = text_similarity(ref_info["REF_PRDT_NM"], rank_names.get(1))
        answer_name_sim = None  # 정답 상품명은 DB에서 조회 필요

        # Miss 유형 분류
        if r1_name_sim is not None and r1_name_sim < 0.3:
            miss_type = "유사실패"
        elif r1_name_sim is not None and r1_name_sim >= 0.3:
            miss_type = "순위밀림(추정)"
        else:
            miss_type = "판단불가"

        row = {
            "REF_STYLE": ref,
            "ITEM": ref_info.get("ITEM", ""),
            "REF_상품명": ref_info["REF_PRDT_NM"],
            "REF_DOMAIN": _clean(ref_info["REF_DOMAIN"]),
            "REF_FIT": _clean(ref_info["REF_FIT"]),
            "REF_가격": ref_info["REF_TAG_PRICE"],
            "정답_스타일": answer,
            "Rank1": rank_styles.get(1, ""),
            "Rank1_상품명": rank_names.get(1, ""),
            "Rank1_DOMAIN": rank_domains.get(1, ""),
            "Rank1_FIT": rank_fits.get(1, ""),
            "Rank1_가격": rank_prices.get(1, ""),
            "Rank2": rank_styles.get(2, ""),
            "Rank2_상품명": rank_names.get(2, ""),
            "Rank3": rank_styles.get(3, ""),
            "Rank3_상품명": rank_names.get(3, ""),
            "REF↔Rank1_상품명유사도": r1_name_sim,
            "REF↔Rank1_DOMAIN일치": "O" if _clean(ref_info["REF_DOMAIN"]) == rank_domains.get(1) else "X",
            "REF↔Rank1_가격차이%": price_diff_pct(ref_info["REF_TAG_PRICE"], rank_prices.get(1)),
            "Miss_유형": miss_type,
        }
        rows.append(row)

    result = pd.DataFrame(rows)
    result = result.sort_values(["ITEM", "REF_STYLE"]).reset_index(drop=True)
    return result


# ── Sheet 3: 정답지_없는_스타일_분석 ──────────────────────────

def build_no_answer_analysis(df):
    """정답지 없는 REF에 대한 프록시 성능 추정."""
    no_answer = df[df["정답지"].isna()].copy()

    # SIM 과다추천 집계 (전체 데이터 기준)
    sim_freq = df[df["RANKING"] == 1]["SIMILAR_STYLE_CD"].value_counts()
    overused_sims = set(sim_freq[sim_freq >= 10].index)

    rows = []
    for ref, group in no_answer.groupby("REF_STYLE"):
        group = group.sort_values("RANKING")
        ref_info = group.iloc[0]

        # Rank 1~3 정보 수집
        domains = []
        fits = []
        prices = []
        names = []
        sim_styles = []
        for _, r in group.iterrows():
            domains.append(_clean(r["SIM_DOMAIN"]))
            fits.append(_clean(r["SIM_FIT"]))
            prices.append(r["SIM_TAG_PRICE"] if pd.notna(r["SIM_TAG_PRICE"]) else None)
            names.append(r["SIM_PRDT_NM"])
            sim_styles.append(r["SIMILAR_STYLE_CD"])

        # 1) 속성 일관성 점수 (Rank 1~3 반환값끼리)
        domain_consistent = len(set(d for d in domains if d is not None)) <= 1
        fit_consistent = len(set(f for f in fits if f is not None)) <= 1
        valid_prices = [p for p in prices if p is not None]
        price_consistent = True
        if len(valid_prices) >= 2:
            price_range = max(valid_prices) - min(valid_prices)
            price_max = max(valid_prices)
            price_consistent = (price_range / price_max * 100) < 30 if price_max > 0 else True

        consistency_score = sum([domain_consistent, fit_consistent, price_consistent])

        # 2) REF ↔ Rank1 상품명 유사도
        name_sim = text_similarity(ref_info["REF_PRDT_NM"], names[0] if names else None)

        # 3) REF ↔ Rank1 가격 편차
        pdiff = price_diff_pct(ref_info["REF_TAG_PRICE"], prices[0] if prices else None)

        # 4) 과다추천 여부
        rank1_style = sim_styles[0] if sim_styles else ""
        is_overused = rank1_style in overused_sims

        # 5) 위험 플래그
        flags = []
        if consistency_score <= 1:
            flags.append("일관성낮음")
        if name_sim is not None and name_sim < 0.3:
            flags.append("상품명불일치")
        if pdiff is not None and pdiff > 40:
            flags.append("가격편차큼")
        if is_overused:
            flags.append("과다추천SIM")

        if len(flags) >= 2:
            risk = "불량추정"
        elif len(flags) == 1:
            risk = "점검필요"
        else:
            risk = "양호"

        rows.append({
            "REF_STYLE": ref,
            "분류": ref_info["분류"],
            "ITEM": ref_info.get("ITEM", ""),
            "REF_상품명": ref_info["REF_PRDT_NM"],
            "REF_DOMAIN": _clean(ref_info["REF_DOMAIN"]),
            "Rank1": rank1_style,
            "Rank1_상품명": names[0] if names else "",
            "상품명유사도": name_sim,
            "가격편차%": pdiff,
            "속성일관성(0~3)": consistency_score,
            "DOMAIN일관": "O" if domain_consistent else "X",
            "FIT일관": "O" if fit_consistent else "X",
            "가격일관": "O" if price_consistent else "X",
            "과다추천SIM": "O" if is_overused else "",
            "위험플래그": risk,
        })

    result = pd.DataFrame(rows)
    result = result.sort_values(["분류", "위험플래그", "ITEM"]).reset_index(drop=True)
    return result


# ── Sheet 4: Wear_vs_ACC_비교 ─────────────────────────────────

def build_category_comparison(df, no_answer_df):
    """Wear vs ACC 구조적 차이 분석."""
    rows = []

    # 전체 데이터 기준 분류별 REF 수
    rank1 = df[df["RANKING"] == 1]
    has_answer_rank1 = df[(df["정답지"].notna()) & (df["RANKING"] == 1)]

    for cat in ["Wear", "ACC"]:
        cat_all = rank1[rank1["분류"] == cat]
        cat_answer = has_answer_rank1[has_answer_rank1["분류"] == cat]
        total = len(cat_all)
        answered = len(cat_answer)

        rows.append({
            "지표": f"── {cat} ──",
            "값": "",
            "비고": "",
        })
        rows.append({
            "지표": "전체 REF 수",
            "값": f"{total}개",
            "비고": "",
        })
        rows.append({
            "지표": "정답지 보유 REF",
            "값": f"{answered}개 ({answered/total*100:.1f}%)" if total > 0 else "0개",
            "비고": "정답지 커버리지",
        })

    # 구분선
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── 속성 가용률 비교 ──", "값": "", "비고": ""})

    # 속성별 null 비율
    for attr, col in [("DOMAIN", "REF_DOMAIN"), ("FIT", "REF_FIT"),
                      ("FAB", "REF_FAB"), ("가격", "REF_TAG_PRICE")]:
        for cat in ["Wear", "ACC"]:
            cat_df = rank1[rank1["분류"] == cat]
            if len(cat_df) == 0:
                continue
            if col in cat_df.columns:
                non_null = cat_df[col].apply(lambda x: _clean(x) is not None).sum()
                rate = non_null / len(cat_df) * 100
                rows.append({
                    "지표": f"{cat} {attr} 가용률",
                    "값": f"{rate:.1f}% ({non_null}/{len(cat_df)})",
                    "비고": "",
                })

    # 구분선
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── 프록시 지표 비교 (정답지 없는 REF) ──", "값": "", "비고": ""})

    # no_answer_df 기반 프록시 지표
    if len(no_answer_df) > 0:
        for cat in ["Wear", "ACC"]:
            cat_na = no_answer_df[no_answer_df["분류"] == cat]
            if len(cat_na) == 0:
                rows.append({
                    "지표": f"{cat} - 해당 없음",
                    "값": "-",
                    "비고": "",
                })
                continue

            avg_sim = cat_na["상품명유사도"].dropna().mean()
            avg_pdiff = cat_na["가격편차%"].dropna().mean()
            avg_consist = cat_na["속성일관성(0~3)"].mean()

            risk_dist = cat_na["위험플래그"].value_counts()
            good = risk_dist.get("양호", 0)
            check = risk_dist.get("점검필요", 0)
            bad = risk_dist.get("불량추정", 0)

            rows.append({
                "지표": f"{cat} 평균 상품명유사도",
                "값": f"{avg_sim:.3f}" if pd.notna(avg_sim) else "-",
                "비고": f"{len(cat_na)}건 대상",
            })
            rows.append({
                "지표": f"{cat} 평균 가격편차%",
                "값": f"{avg_pdiff:.1f}%" if pd.notna(avg_pdiff) else "-",
                "비고": "",
            })
            rows.append({
                "지표": f"{cat} 평균 속성일관성",
                "값": f"{avg_consist:.2f}/3",
                "비고": "",
            })
            rows.append({
                "지표": f"{cat} 위험분포",
                "값": f"양호 {good} / 점검필요 {check} / 불량추정 {bad}",
                "비고": f"불량률 {bad/len(cat_na)*100:.1f}%",
            })

    # ACC 개선 방향
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── ACC 모델링 개선 방향 ──", "값": "", "비고": ""})

    acc_rank1 = rank1[rank1["분류"] == "ACC"]
    if len(acc_rank1) > 0:
        # ACC에서 부족한 속성 식별
        for attr, col in [("DOMAIN", "REF_DOMAIN"), ("FIT", "REF_FIT"), ("FAB", "REF_FAB")]:
            non_null = acc_rank1[col].apply(lambda x: _clean(x) is not None).sum()
            rate = non_null / len(acc_rank1) * 100
            if rate < 50:
                rows.append({
                    "지표": f"  {attr} 부족",
                    "값": f"가용률 {rate:.0f}%",
                    "비고": f"→ {attr} 속성 보강 또는 가중치 조정 필요",
                })

        rows.append({
            "지표": "  정답지 수집",
            "값": "0건 (미확보)",
            "비고": "→ ACC 카테고리 정답지 수집이 최우선",
        })

    return pd.DataFrame(rows)


# ── Sheet 5: 최적화_방향_제안 ──────────────────────────────────

def build_optimization_suggestions(df, miss_df, perf_summary_df):
    """ITEM별 성능 격차와 miss 패턴 기반 모델 개선 방향."""
    has_answer = df[df["정답지"].notna()].copy()
    rows = []

    # ── 1. ITEM별 가중치 차별화 제안
    rows.append({"주제": "1. ITEM별 가중치 차별화 제안", "내용": "", "근거": ""})
    rows.append({"주제": "", "내용": "", "근거": ""})

    # ITEM별 성능 (perf_summary_df에서 추출)
    for _, r in perf_summary_df.iterrows():
        label = r["구분"]
        if label == "전체":
            continue
        grade = r.get("등급", "")
        if grade == "A":
            rows.append({
                "주제": f"  {label}",
                "내용": f"현 가중치 유지 (등급 {grade})",
                "근거": f"Hit@1={r['Hit@1']}, Miss={r['Miss']}",
            })
        elif grade in ("C", "D"):
            rows.append({
                "주제": f"  {label}",
                "내용": f"가중치 재조정 필요 (등급 {grade})",
                "근거": f"Hit@1={r['Hit@1']}, Miss={r['Miss']}",
            })
        elif grade == "B":
            rows.append({
                "주제": f"  {label}",
                "내용": f"소폭 개선 여지 (등급 {grade})",
                "근거": f"Hit@1={r['Hit@1']}, Miss={r['Miss']}",
            })

    # ── 2. 속성별 조정 방향 (Miss 케이스 분석)
    rows.append({"주제": "", "내용": "", "근거": ""})
    rows.append({"주제": "2. 속성별 조정 방향 (Miss 패턴 분석)", "내용": "", "근거": ""})
    rows.append({"주제": "", "내용": "", "근거": ""})

    if len(miss_df) > 0:
        # DOMAIN 불일치율
        domain_mismatch = (miss_df["REF↔Rank1_DOMAIN일치"] == "X").sum()
        domain_total = len(miss_df)
        rows.append({
            "주제": "  DOMAIN",
            "내용": f"Miss 중 DOMAIN 불일치: {domain_mismatch}/{domain_total} ({domain_mismatch/domain_total*100:.0f}%)",
            "근거": "DOMAIN 불일치가 높으면 → DOMAIN 가중치 상향 또는 하드필터 강화",
        })

        # 상품명 유사도 분포
        name_sims = miss_df["REF↔Rank1_상품명유사도"].dropna()
        if len(name_sims) > 0:
            high_sim = (name_sims >= 0.5).sum()
            low_sim = (name_sims < 0.3).sum()
            rows.append({
                "주제": "  상품명",
                "내용": f"Miss 중 상품명유사도≥0.5: {high_sim}건(순위밀림), <0.3: {low_sim}건(완전실패)",
                "근거": "유사도 높은데 miss → 다른 속성(DOMAIN/FIT)이 원인",
            })

        # 가격 편차
        pdiffs = miss_df["REF↔Rank1_가격차이%"].dropna()
        if len(pdiffs) > 0:
            high_pdiff = (pdiffs >= 30).sum()
            rows.append({
                "주제": "  가격",
                "내용": f"Miss 중 가격차이≥30%: {high_pdiff}/{len(pdiffs)}건",
                "근거": "가격 편차가 크면 → 가격 유사도 가중치 상향 검토",
            })

        # Miss 유형 분포
        miss_type_dist = miss_df["Miss_유형"].value_counts()
        rows.append({
            "주제": "  Miss 유형 분포",
            "내용": " / ".join(f"{t}: {c}건" for t, c in miss_type_dist.items()),
            "근거": "유사실패 → 모델 자체 한계, 순위밀림 → K 확장 또는 가중치 미세 조정",
        })

    # ── 3. 과다추천 방지 제안
    rows.append({"주제": "", "내용": "", "근거": ""})
    rows.append({"주제": "3. 과다추천 방지 제안", "내용": "", "근거": ""})
    rows.append({"주제": "", "내용": "", "근거": ""})

    sim_freq = df[df["RANKING"] == 1]["SIMILAR_STYLE_CD"].value_counts()
    overused = sim_freq[sim_freq >= 10]
    if len(overused) > 0:
        rows.append({
            "주제": "  현황",
            "내용": f"{len(overused)}개 SIM이 10회 이상 Rank1으로 반환",
            "근거": f"최다: {overused.index[0]} ({overused.iloc[0]}회)",
        })
        rows.append({
            "주제": "  제안1",
            "내용": "Top-K 반환 시 diversity penalty 도입",
            "근거": "같은 SIM이 N회 이상 반환되면 score 차감",
        })
        rows.append({
            "주제": "  제안2",
            "내용": "SIM pool에서 이미 많이 매칭된 스타일의 점수 감소",
            "근거": "전체 추천 다양성 확보",
        })
    else:
        rows.append({
            "주제": "  현황",
            "내용": "심각한 과다추천 없음 (10회 이상 반복 SIM 없음)",
            "근거": "",
        })

    # ── 4. 정답지 확대 제안
    rows.append({"주제": "", "내용": "", "근거": ""})
    rows.append({"주제": "4. 정답지 확대 제안", "내용": "", "근거": ""})
    rows.append({"주제": "", "내용": "", "근거": ""})

    rows.append({
        "주제": "  ACC 카테고리",
        "내용": "정답지 0건 → ACC 정답지 수집 최우선",
        "근거": "ACC 87개 REF의 성능을 측정할 수 없는 상태",
    })

    # Miss 많은 ITEM
    if len(perf_summary_df) > 0:
        weak_items = []
        for _, r in perf_summary_df.iterrows():
            if r["구분"] == "전체" or r.get("등급", "") in ("", "-", "A", "B"):
                continue
            weak_items.append(r["구분"])
        if weak_items:
            rows.append({
                "주제": "  약점 ITEM",
                "내용": f"추가 정답지 확보 필요: {', '.join(weak_items)}",
                "근거": "Hit@1 낮은 ITEM의 정답지를 늘려 miss 패턴 분석 강화",
            })

    return pd.DataFrame(rows)


# ── Sheet 6: 하드필터_위반_검출 ────────────────────────────────

def build_hardfilter_violations(df):
    """성별 위반, 가격 극단, DOMAIN 교차 등 모델 버그 검출."""
    violations = []

    for _, row in df.iterrows():
        issues = []

        # 1) 성별 위반: 남성↔여성 (공용 없이)
        if gender_compatible(row["REF_SEX_NM"], row["SIM_SEX_NM"]) is False:
            issues.append(
                f"성별 위반: {_clean(row['REF_SEX_NM'])}→{_clean(row['SIM_SEX_NM'])}"
            )

        # 2) 극단적 가격 차이 (50% 이상)
        pdiff = price_diff_pct(row["REF_TAG_PRICE"], row["SIM_TAG_PRICE"])
        if pdiff is not None and pdiff >= 50:
            ref_p = row["REF_TAG_PRICE"]
            sim_p = row["SIM_TAG_PRICE"]
            issues.append(f"가격 차이 극단: {pdiff:.0f}% ({ref_p:,.0f}→{sim_p:,.0f})")

        # 3) DOMAIN 교차 매칭 중 문제되는 패턴
        ref_d = _clean(row["REF_DOMAIN"])
        sim_d = _clean(row["SIM_DOMAIN"])
        if ref_d and sim_d and ref_d != sim_d:
            problematic_pairs = [
                frozenset(["Monogram", "Varsity"]),
                frozenset(["Megagram", "Basic"]),
            ]
            if frozenset([ref_d, sim_d]) in problematic_pairs:
                issues.append(f"DOMAIN 교차: {ref_d}→{sim_d}")

        if issues:
            violations.append({
                "Rank": int(row["RANKING"]),
                "분류": row["분류"],
                "ITEM": row.get("ITEM", ""),
                "REF_STYLE": row["REF_STYLE"],
                "REF_상품명": row["REF_PRDT_NM"],
                "REF_성별": _clean(row["REF_SEX_NM"]),
                "REF_가격": row["REF_TAG_PRICE"],
                "REF_DOMAIN": _clean(row["REF_DOMAIN"]),
                "SIM_STYLE": row["SIMILAR_STYLE_CD"],
                "SIM_상품명": row["SIM_PRDT_NM"],
                "SIM_성별": _clean(row["SIM_SEX_NM"]),
                "SIM_가격": row["SIM_TAG_PRICE"],
                "SIM_DOMAIN": _clean(row["SIM_DOMAIN"]),
                "위반내용": " | ".join(issues),
                "심각도": "치명" if any("성별" in i for i in issues) else "주의",
            })

    result = pd.DataFrame(violations)
    if len(result) > 0:
        result = result.sort_values(
            ["심각도", "분류", "REF_STYLE"], ascending=[True, True, True]
        ).reset_index(drop=True)
    return result


# ── Sheet 7: Rank_변별력_분석 ──────────────────────────────────

def build_rank_discrimination(df):
    """Rank 1이 정말 Rank 3보다 나은지 + 역전 검출."""
    # Rank별 평균 지표 요약
    summary_rows = []
    for rank in [1, 2, 3]:
        rank_df = df[df["RANKING"] == rank]

        domain_match = sum(
            1 for _, r in rank_df.iterrows()
            if _clean(r["REF_DOMAIN"]) == _clean(r["SIM_DOMAIN"])
            and _clean(r["REF_DOMAIN"]) is not None
        )
        domain_total = sum(
            1 for _, r in rank_df.iterrows()
            if _clean(r["REF_DOMAIN"]) is not None
            and _clean(r["SIM_DOMAIN"]) is not None
        )
        domain_rate = domain_match / domain_total * 100 if domain_total > 0 else 0

        name_sims = [
            text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
            for _, r in rank_df.iterrows()
        ]
        name_sims = [x for x in name_sims if x is not None]
        avg_name = sum(name_sims) / len(name_sims) if name_sims else 0

        pdiffs = [
            price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"])
            for _, r in rank_df.iterrows()
        ]
        pdiffs = [x for x in pdiffs if x is not None]
        avg_pdiff = sum(pdiffs) / len(pdiffs) if pdiffs else 0

        summary_rows.append({
            "Rank": f"Rank {rank}",
            "건수": len(rank_df),
            "DOMAIN 일치율": f"{domain_rate:.1f}%",
            "평균 상품명유사도": f"{avg_name:.3f}",
            "평균 가격차이%": f"{avg_pdiff:.1f}%",
        })

    # REF별 Rank 역전 분석
    detail_rows = []
    for ref, group in df.groupby("REF_STYLE"):
        group = group.sort_values("RANKING")
        ref_info = group.iloc[0]

        sims = {}
        for _, row in group.iterrows():
            rank = int(row["RANKING"])
            domain_ok = 1 if _clean(row["REF_DOMAIN"]) == _clean(row["SIM_DOMAIN"]) else 0
            name_sim = text_similarity(row["REF_PRDT_NM"], row["SIM_PRDT_NM"])
            pdiff = price_diff_pct(row["REF_TAG_PRICE"], row["SIM_TAG_PRICE"])
            sims[rank] = {
                "style": row["SIMILAR_STYLE_CD"],
                "name": row["SIM_PRDT_NM"],
                "domain_ok": domain_ok,
                "name_sim": name_sim or 0,
                "price_diff": pdiff or 0,
            }

        row_data = {
            "REF_STYLE": ref,
            "분류": ref_info["분류"],
            "ITEM": ref_info.get("ITEM", ""),
        }

        for r in [1, 2, 3]:
            if r in sims:
                row_data[f"Rank{r}_DOMAIN일치"] = "O" if sims[r]["domain_ok"] else "X"
                row_data[f"Rank{r}_상품명유사도"] = sims[r]["name_sim"]
                row_data[f"Rank{r}_가격차이%"] = sims[r]["price_diff"]

        # Rank 1 vs 3 역전 판정
        if 1 in sims and 3 in sims:
            r1, r3 = sims[1], sims[3]
            r3_better = (
                r3["name_sim"] > r1["name_sim"] + 0.1
                and r3["price_diff"] < r1["price_diff"]
            )
            no_diff = abs(r1["name_sim"] - r3["name_sim"]) < 0.05
            if r3_better:
                row_data["Rank역전"] = "Rank3>Rank1"
            elif no_diff:
                row_data["Rank역전"] = "차이없음"
            else:
                row_data["Rank역전"] = ""
        else:
            row_data["Rank역전"] = ""

        detail_rows.append(row_data)

    summary_df = pd.DataFrame(summary_rows)
    detail_df = pd.DataFrame(detail_rows)

    # 역전 통계
    if len(detail_df) > 0 and "Rank역전" in detail_df.columns:
        inv = (detail_df["Rank역전"] == "Rank3>Rank1").sum()
        nodiff = (detail_df["Rank역전"] == "차이없음").sum()
        total = len(detail_df)
        summary_rows.append({
            "Rank": "역전(3>1)",
            "건수": f"{inv}/{total} ({inv/total*100:.1f}%)",
            "DOMAIN 일치율": "",
            "평균 상품명유사도": "",
            "평균 가격차이%": "",
        })
        summary_rows.append({
            "Rank": "차이없음",
            "건수": f"{nodiff}/{total} ({nodiff/total*100:.1f}%)",
            "DOMAIN 일치율": "",
            "평균 상품명유사도": "",
            "평균 가격차이%": "",
        })
        summary_df = pd.DataFrame(summary_rows)

    return summary_df, detail_df


# ── Sheet 8: 속성별_일치율_통계 ────────────────────────────────

def build_attribute_stats(df):
    """DOMAIN/FIT/성별/가격 전체 일치율 기초 통계."""
    total = len(df)
    rows = []

    # DOMAIN 일치
    domain_match = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_DOMAIN"]) == _clean(r["SIM_DOMAIN"])
        and _clean(r["REF_DOMAIN"]) is not None
    )
    domain_total = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_DOMAIN"]) is not None
        and _clean(r["SIM_DOMAIN"]) is not None
    )
    rows.append({
        "속성": "DOMAIN 일치율",
        "값": f"{domain_match/domain_total*100:.1f}%" if domain_total else "-",
        "건수": f"{domain_match}/{domain_total}",
        "비고": "Exact match 기준",
    })

    # FIT 일치
    fit_match = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_FIT"]) is not None
        and _clean(r["SIM_FIT"]) is not None
        and _clean(r["REF_FIT"]) == _clean(r["SIM_FIT"])
    )
    fit_total = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_FIT"]) is not None
        and _clean(r["SIM_FIT"]) is not None
    )
    rows.append({
        "속성": "FIT 일치율",
        "값": f"{fit_match/fit_total*100:.1f}%" if fit_total else "-",
        "건수": f"{fit_match}/{fit_total}",
        "비고": "값 존재하는 건만 대상",
    })

    # 성별 호환
    gender_ok = sum(
        1 for _, r in df.iterrows()
        if gender_compatible(r["REF_SEX_NM"], r["SIM_SEX_NM"]) is True
    )
    rows.append({
        "속성": "성별 호환율",
        "값": f"{gender_ok/total*100:.1f}%",
        "건수": f"{gender_ok}/{total}",
        "비고": "공용↔남녀 허용",
    })

    # 가격 평균 차이
    pdiffs = [
        price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"])
        for _, r in df.iterrows()
    ]
    pdiffs = [x for x in pdiffs if x is not None]
    rows.append({
        "속성": "가격 평균차이",
        "값": f"{sum(pdiffs)/len(pdiffs):.1f}%" if pdiffs else "-",
        "건수": f"{len(pdiffs)}건",
        "비고": "max(REF,SIM) 대비 차이율",
    })

    # 상품명 평균 유사도
    nsims = [
        text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
        for _, r in df.iterrows()
    ]
    nsims = [x for x in nsims if x is not None]
    rows.append({
        "속성": "상품명 평균유사도",
        "값": f"{sum(nsims)/len(nsims):.3f}" if nsims else "-",
        "건수": f"{len(nsims)}건",
        "비고": "SequenceMatcher 기준",
    })

    # Wear vs ACC 분류별
    for cat in ["Wear", "ACC"]:
        cat_df = df[df["분류"] == cat]
        if len(cat_df) == 0:
            continue

        rows.append({"속성": f"── {cat} ──", "값": "", "건수": f"{len(cat_df)}건", "비고": ""})

        cat_nsims = [
            text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
            for _, r in cat_df.iterrows()
        ]
        cat_nsims = [x for x in cat_nsims if x is not None]
        rows.append({
            "속성": "  상품명 평균유사도",
            "값": f"{sum(cat_nsims)/len(cat_nsims):.3f}" if cat_nsims else "-",
            "건수": "", "비고": "",
        })

        cat_pdiffs = [
            price_diff_pct(r["REF_TAG_PRICE"], r["SIM_TAG_PRICE"])
            for _, r in cat_df.iterrows()
        ]
        cat_pdiffs = [x for x in cat_pdiffs if x is not None]
        rows.append({
            "속성": "  가격 평균차이",
            "값": f"{sum(cat_pdiffs)/len(cat_pdiffs):.1f}%" if cat_pdiffs else "-",
            "건수": "", "비고": "",
        })

        cat_domain = sum(
            1 for _, r in cat_df.iterrows()
            if _clean(r["REF_DOMAIN"]) == _clean(r["SIM_DOMAIN"])
            and _clean(r["REF_DOMAIN"]) is not None
        )
        cat_domain_t = sum(
            1 for _, r in cat_df.iterrows()
            if _clean(r["REF_DOMAIN"]) is not None
            and _clean(r["SIM_DOMAIN"]) is not None
        )
        rows.append({
            "속성": "  DOMAIN 일치율",
            "값": f"{cat_domain/cat_domain_t*100:.1f}%" if cat_domain_t else "-",
            "건수": f"{cat_domain}/{cat_domain_t}",
            "비고": "",
        })

    return pd.DataFrame(rows)


# ── Sheet 9: 서브시즌_크로스매칭 ──────────────────────────────

def build_subseason_cross(df):
    """의류 Rank1 서브시즌 일치/불일치 현황 및 정답 교차분석."""
    wear = df[(df["분류"] == "Wear") & (df["RANKING"] == 1)].copy()
    total = len(wear)
    if total == 0:
        return pd.DataFrame()

    same = (wear["REF_SUB"] == wear["SIM_SUB"]).sum()
    diff = total - same

    rows = []
    rows.append({"지표": "── 전체 현황 ──", "값": "", "비고": ""})
    rows.append({
        "지표": "Wear Rank1 전체",
        "값": f"{total}건",
        "비고": "",
    })
    rows.append({
        "지표": "서브시즌 일치",
        "값": f"{same}건 ({same / total * 100:.1f}%)",
        "비고": "REF 끝자리 == SIM 끝자리",
    })
    rows.append({
        "지표": "서브시즌 불일치 (크로스매칭)",
        "값": f"{diff}건 ({diff / total * 100:.1f}%)",
        "비고": "",
    })

    # ITEM별 크로스매칭률
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── ITEM별 크로스매칭률 ──", "값": "", "비고": ""})
    for item, g in wear.groupby("ITEM"):
        if pd.isna(item):
            continue
        n = len(g)
        cross = (g["REF_SUB"] != g["SIM_SUB"]).sum()
        rows.append({
            "지표": f"  {item}",
            "값": f"{cross}/{n} ({cross / n * 100:.0f}%)",
            "비고": f"n={n}",
        })

    # 서브시즌 × 정답 교차분석 (정답지 있는 건만)
    has_answer = wear[wear["채점"].notna()].copy()
    if len(has_answer) > 0:
        rows.append({"지표": "", "값": "", "비고": ""})
        rows.append({"지표": "── 서브시즌 × 정답 교차분석 ──", "값": "", "비고": ""})
        same_mask = has_answer["REF_SUB"] == has_answer["SIM_SUB"]
        for label, mask in [("서브시즌 일치", same_mask), ("서브시즌 불일치", ~same_mask)]:
            subset = has_answer[mask]
            n = len(subset)
            if n == 0:
                continue
            correct = (subset["채점"] == "O").sum()
            rows.append({
                "지표": f"  {label}",
                "값": f"정답률 {correct}/{n} ({correct / n * 100:.1f}%)",
                "비고": "",
            })

    return pd.DataFrame(rows)


# ── Sheet 10: 서브시즌_시뮬레이션 ────────────────────────────

def build_subseason_simulation(df):
    """AS-IS vs TO-BE(서브시즌 하드필터) 성능 비교."""
    wear = df[df["분류"] == "Wear"].copy()
    has_answer = wear[wear["정답지"].notna()].copy()
    if len(has_answer) == 0:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    def _calc_hit(data):
        """REF별 Hit@K 계산. returns dict {ref: first_hit_rank}."""
        hit_at = {}
        for ref, g in data.groupby("REF_STYLE"):
            g = g.sort_values("RANKING")
            for _, r in g.iterrows():
                if r["채점"] == "O":
                    hit_at[ref] = int(r["RANKING"])
                    break
        return hit_at

    # ── AS-IS 성능 (전체) ──
    asis_hit = _calc_hit(has_answer)
    refs = has_answer[has_answer["RANKING"] == 1]["REF_STYLE"].unique()
    n_total = len(refs)
    asis_h1 = sum(1 for v in asis_hit.values() if v <= 1)
    asis_h3 = sum(1 for v in asis_hit.values() if v <= 3)
    asis_mrr = sum(1.0 / v for v in asis_hit.values()) / n_total if n_total > 0 else 0

    # ── TO-BE: 서브시즌 일치만 남기고 re-rank ──
    filtered_rows = []
    for ref, g in has_answer.groupby("REF_STYLE"):
        ref_sub = g["REF_SUB"].iloc[0]
        same_sub = g[g["SIM_SUB"] == ref_sub].copy()
        if len(same_sub) == 0:
            continue
        same_sub = same_sub.sort_values("RANKING")
        same_sub["RANKING"] = range(1, len(same_sub) + 1)
        filtered_rows.append(same_sub)

    tobe_h1, tobe_h3, tobe_mrr, n_tobe = 0, 0, 0, 0
    if filtered_rows:
        filtered_df = pd.concat(filtered_rows)
        tobe_hit = _calc_hit(filtered_df)
        tobe_refs = filtered_df[filtered_df["RANKING"] == 1]["REF_STYLE"].unique()
        n_tobe = len(tobe_refs)
        tobe_h1 = sum(1 for v in tobe_hit.values() if v <= 1)
        tobe_h3 = sum(1 for v in tobe_hit.values() if v <= 3)
        tobe_mrr = sum(1.0 / v for v in tobe_hit.values()) / n_tobe if n_tobe > 0 else 0

    # 요약 테이블
    summary_rows = [
        {
            "구분": "AS-IS (현재)",
            "대상": f"{n_total}건",
            "Hit@1": f"{asis_h1}/{n_total} ({asis_h1 / n_total * 100:.1f}%)" if n_total else "-",
            "Hit@3": f"{asis_h3}/{n_total} ({asis_h3 / n_total * 100:.1f}%)" if n_total else "-",
            "MRR": f"{asis_mrr:.3f}",
        },
        {
            "구분": "TO-BE (서브시즌 하드필터)",
            "대상": f"{n_tobe}건",
            "Hit@1": f"{tobe_h1}/{n_tobe} ({tobe_h1 / n_tobe * 100:.1f}%)" if n_tobe else "-",
            "Hit@3": f"{tobe_h3}/{n_tobe} ({tobe_h3 / n_tobe * 100:.1f}%)" if n_tobe else "-",
            "MRR": f"{tobe_mrr:.3f}",
        },
        {
            "구분": "개선효과",
            "대상": "",
            "Hit@1": f"+{tobe_h1 - asis_h1}건",
            "Hit@3": f"+{tobe_h3 - asis_h3}건",
            "MRR": f"+{tobe_mrr - asis_mrr:.3f}",
        },
    ]
    summary_df = pd.DataFrame(summary_rows)

    # ── ITEM별 시뮬레이션 ──
    def _grade(pct):
        if pct >= 70:
            return "A"
        if pct >= 50:
            return "B"
        if pct >= 30:
            return "C"
        return "D"

    item_rows = []
    for item, item_df in has_answer.groupby("ITEM"):
        if pd.isna(item):
            continue
        item_refs = item_df[item_df["RANKING"] == 1]["REF_STYLE"].unique()
        n = len(item_refs)
        if n < 2:
            continue

        # AS-IS
        a_hit = _calc_hit(item_df)
        a_h1 = sum(1 for v in a_hit.values() if v <= 1)

        # TO-BE
        t_parts = []
        for ref in item_refs:
            g = item_df[item_df["REF_STYLE"] == ref].sort_values("RANKING")
            ref_sub = g["REF_SUB"].iloc[0]
            ss = g[g["SIM_SUB"] == ref_sub].copy()
            if len(ss) == 0:
                continue
            ss["RANKING"] = range(1, len(ss) + 1)
            t_parts.append(ss)
        t_h1 = 0
        if t_parts:
            t_hit = _calc_hit(pd.concat(t_parts))
            t_h1 = sum(1 for v in t_hit.values() if v <= 1)

        rank1 = item_df[item_df["RANKING"] == 1]
        cross_rate = (rank1["REF_SUB"] != rank1["SIM_SUB"]).sum() / n * 100

        a_pct = a_h1 / n * 100
        t_pct = t_h1 / n * 100
        g_a = _grade(a_pct)
        g_t = _grade(t_pct)

        item_rows.append({
            "ITEM": item,
            "n": n,
            "크로스매칭률": f"{cross_rate:.0f}%",
            "AS-IS_Hit@1": f"{a_h1}/{n} ({a_pct:.0f}%)",
            "TO-BE_Hit@1": f"{t_h1}/{n} ({t_pct:.0f}%)",
            "delta": f"+{t_h1 - a_h1}" if t_h1 - a_h1 > 0 else str(t_h1 - a_h1),
            "등급변화": f"{g_a}→{g_t}" if g_a != g_t else g_a,
        })
    item_rows.sort(key=lambda x: -int(x["delta"].replace("+", "")))
    item_df_result = pd.DataFrame(item_rows)

    # ── Hit@1 신규 획득 케이스 ──
    improvements = []
    if filtered_rows:
        for ref, g in has_answer.groupby("REF_STYLE"):
            g = g.sort_values("RANKING")
            rank1_row = g[g["RANKING"] == 1].iloc[0]
            asis_ok = rank1_row["채점"] == "O"
            if asis_ok:
                continue

            ref_sub = g["REF_SUB"].iloc[0]
            ss = g[g["SIM_SUB"] == ref_sub].sort_values("RANKING")
            if len(ss) == 0:
                continue
            first = ss.iloc[0]
            if first["채점"] == "O":
                improvements.append({
                    "REF_STYLE": ref,
                    "ITEM": rank1_row.get("ITEM", ""),
                    "REF_상품명": rank1_row["REF_PRDT_NM"],
                    "AS-IS_Rank1": rank1_row["SIMILAR_STYLE_CD"],
                    "AS-IS_Rank1_SUB": rank1_row["SIM_SUB"],
                    "AS-IS_상품명": rank1_row["SIM_PRDT_NM"],
                    "TO-BE_Rank1": first["SIMILAR_STYLE_CD"],
                    "TO-BE_Rank1_SUB": first["SIM_SUB"],
                    "TO-BE_상품명": first["SIM_PRDT_NM"],
                    "기존Rank": int(first["RANKING"]) if "RANKING" in first.index else "",
                })
    improve_df = pd.DataFrame(improvements)

    return summary_df, item_df_result, improve_df


# ── Sheet 11: DOMAIN2_분석 ───────────────────────────────────

def build_domain2_analysis(df):
    """DOMAIN1×DOMAIN2 교차표 및 불일치 분석 (Wear)."""
    wear_r1 = df[(df["분류"] == "Wear") & (df["RANKING"] == 1)].copy()
    rows = []

    # DOMAIN1 × DOMAIN2 교차표 (REF 기준)
    d1d2 = wear_r1[wear_r1["REF_DOMAIN2"].notna()].copy()
    if len(d1d2) > 0:
        ct = pd.crosstab(d1d2["REF_DOMAIN"], d1d2["REF_DOMAIN2"])
        rows.append({"지표": "── DOMAIN1 × DOMAIN2 교차표 (Wear REF) ──", "값": "", "비고": ""})
        for d1 in ct.index:
            d2_vals = []
            for d2 in ct.columns:
                if ct.loc[d1, d2] > 0:
                    d2_vals.append(f"{d2}({ct.loc[d1, d2]})")
            rows.append({
                "지표": f"  {d1}",
                "값": ", ".join(d2_vals),
                "비고": f"합계 {ct.loc[d1].sum()}건",
            })

    # Rank1에서 DOMAIN1 일치 + DOMAIN2 불일치 현황
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── DOMAIN1 일치 + DOMAIN2 불일치 현황 ──", "값": "", "비고": ""})

    d1_match = wear_r1[
        (wear_r1["REF_DOMAIN"].notna())
        & (wear_r1["SIM_DOMAIN"].notna())
        & (wear_r1["REF_DOMAIN"] == wear_r1["SIM_DOMAIN"])
    ].copy()

    if len(d1_match) > 0:
        both_d2 = d1_match[
            (d1_match["REF_DOMAIN2"].notna()) & (d1_match["SIM_DOMAIN2"].notna())
        ]
        d2_match = (both_d2["REF_DOMAIN2"] == both_d2["SIM_DOMAIN2"]).sum()
        d2_mismatch = len(both_d2) - d2_match
        rows.append({
            "지표": "DOMAIN1 일치 건 중 DOMAIN2 보유",
            "값": f"{len(both_d2)}건",
            "비고": f"(DOMAIN1 일치 전체: {len(d1_match)}건)",
        })
        rows.append({
            "지표": "  DOMAIN2도 일치",
            "값": f"{d2_match}건 ({d2_match / len(both_d2) * 100:.1f}%)" if len(both_d2) else "-",
            "비고": "",
        })
        rows.append({
            "지표": "  DOMAIN2 불일치",
            "값": f"{d2_mismatch}건 ({d2_mismatch / len(both_d2) * 100:.1f}%)" if len(both_d2) else "-",
            "비고": "",
        })

        # 불일치 패턴
        if d2_mismatch > 0:
            mismatch_df = both_d2[both_d2["REF_DOMAIN2"] != both_d2["SIM_DOMAIN2"]]
            pairs = mismatch_df.groupby(["REF_DOMAIN2", "SIM_DOMAIN2"]).size().sort_values(ascending=False)
            rows.append({"지표": "", "값": "", "비고": ""})
            rows.append({"지표": "  불일치 패턴 (REF_D2→SIM_D2)", "값": "", "비고": ""})
            for (rd2, sd2), cnt in pairs.head(10).items():
                rows.append({
                    "지표": f"    {rd2} → {sd2}",
                    "값": f"{cnt}건",
                    "비고": "",
                })

    # DOMAIN2 × 정답 교차분석 (정답지 있는 건만)
    has_answer = wear_r1[wear_r1["채점"].notna()].copy()
    if len(has_answer) > 0:
        both_d2_ans = has_answer[
            (has_answer["REF_DOMAIN2"].notna()) & (has_answer["SIM_DOMAIN2"].notna())
        ]
        if len(both_d2_ans) > 0:
            rows.append({"지표": "", "값": "", "비고": ""})
            rows.append({"지표": "── DOMAIN2 × 정답 교차분석 ──", "값": "", "비고": ""})
            d2_same = both_d2_ans["REF_DOMAIN2"] == both_d2_ans["SIM_DOMAIN2"]
            for label, mask in [("DOMAIN2 일치", d2_same), ("DOMAIN2 불일치", ~d2_same)]:
                subset = both_d2_ans[mask]
                n = len(subset)
                if n == 0:
                    continue
                correct = (subset["채점"] == "O").sum()
                rows.append({
                    "지표": f"  {label}",
                    "값": f"정답률 {correct}/{n} ({correct / n * 100:.1f}%)",
                    "비고": "",
                })

    return pd.DataFrame(rows)


# ── Sheet 12: 용품_상세분석 ──────────────────────────────────

def build_acc_detail(df):
    """ACC 넌시즌 매칭 현황, ITEM별 현황, 과다추천 상세."""
    acc = df[(df["분류"] == "ACC") & (df["RANKING"] == 1)].copy()
    rows = []
    if len(acc) == 0:
        return pd.DataFrame()

    # ── 넌시즌 매칭 현황 ──
    rows.append({"지표": "── 넌시즌 매칭 현황 ──", "값": "", "비고": ""})
    nonsesn = acc[acc["REF_SESN"] == "논시즌"]
    n_nonsesn = len(nonsesn)
    rows.append({
        "지표": "ACC 논시즌 REF",
        "값": f"{n_nonsesn}건 (전체 ACC {len(acc)}건 중 {n_nonsesn / len(acc) * 100:.1f}%)",
        "비고": "",
    })

    if n_nonsesn > 0:
        ns_to_ns = (nonsesn["SIM_SESN"] == "논시즌").sum()
        ns_to_s = n_nonsesn - ns_to_ns
        rows.append({
            "지표": "  논시즌 → 논시즌 매칭",
            "값": f"{ns_to_ns}건 ({ns_to_ns / n_nonsesn * 100:.1f}%)",
            "비고": "",
        })
        rows.append({
            "지표": "  논시즌 → 시즌 매칭",
            "값": f"{ns_to_s}건 ({ns_to_s / n_nonsesn * 100:.1f}%)",
            "비고": "시즌 상품을 논시즌에 매칭",
        })

    # ── ACC ITEM별 현황 ──
    rows.append({"지표": "", "값": "", "비고": ""})
    rows.append({"지표": "── ACC ITEM별 현황 ──", "값": "", "비고": ""})

    # REF의 FIT 가용률과 FIT 일치율을 ITEM별로
    all_acc = df[(df["분류"] == "ACC") & (df["RANKING"] == 1)]
    sim_freq = df[df["RANKING"] == 1]["SIMILAR_STYLE_CD"].value_counts()
    overused_sims = set(sim_freq[sim_freq >= 10].index)

    for item, g in all_acc.groupby("ITEM"):
        if pd.isna(item):
            continue
        n = len(g)
        n_nonsesn_item = (g["REF_SESN"] == "논시즌").sum()

        # FIT 가용률
        fit_avail = g["REF_FIT"].apply(lambda x: _clean(x) is not None).sum()
        fit_rate = fit_avail / n * 100 if n > 0 else 0

        # FIT 일치율 (값 있는 건만)
        both_fit = g[(g["REF_FIT"].apply(lambda x: _clean(x) is not None))
                     & (g["SIM_FIT"].apply(lambda x: _clean(x) is not None))]
        fit_match = sum(
            1 for _, r in both_fit.iterrows()
            if _clean(r["REF_FIT"]) == _clean(r["SIM_FIT"])
        )
        fit_match_rate = fit_match / len(both_fit) * 100 if len(both_fit) else 0

        # 과다추천 SIM 수
        overused_cnt = g["SIMILAR_STYLE_CD"].isin(overused_sims).sum()

        rows.append({
            "지표": f"  {item}",
            "값": (f"n={n}, 논시즌={n_nonsesn_item}, "
                   f"FIT가용={fit_rate:.0f}%, FIT일치={fit_match_rate:.0f}%"),
            "비고": f"과다추천SIM={overused_cnt}건" if overused_cnt > 0 else "",
        })

    # ── 신발(CV) 과다추천 상세 ──
    cv_acc = df[(df["분류"] == "ACC") & (df["RANKING"] == 1) & (df["ITEM"] == "CV")]
    if len(cv_acc) > 0:
        cv_overused = cv_acc[cv_acc["SIMILAR_STYLE_CD"].isin(overused_sims)]
        if len(cv_overused) > 0:
            rows.append({"지표": "", "값": "", "비고": ""})
            rows.append({"지표": "── 신발(CV) 과다추천 상세 ──", "값": "", "비고": ""})
            for sim, cnt in cv_overused["SIMILAR_STYLE_CD"].value_counts().head(5).items():
                sim_nm = cv_overused[cv_overused["SIMILAR_STYLE_CD"] == sim].iloc[0]["SIM_PRDT_NM"]
                total_use = sim_freq.get(sim, 0)
                rows.append({
                    "지표": f"  {sim}",
                    "값": f"CV내 {cnt}회 (전체 {total_use}회)",
                    "비고": f"{sim_nm}",
                })

    return pd.DataFrame(rows)


# ── Excel formatting ───────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = CELL_FONT


def write_df(ws, df, start_row=1):
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, float) and pd.isna(val):
                ws.cell(row=r_idx, column=c_idx, value="")
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)


# ── main ───────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("STEP 0 유사스타일 모델 성능 측정 리포트")
    print("=" * 60)

    print("\n[1/13] 데이터 로딩...")
    df, db = load_data()
    rank1 = df[df["RANKING"] == 1]
    has_answer_rank1 = df[(df["정답지"].notna()) & (df["RANKING"] == 1)]
    no_answer_rank1 = df[(df["정답지"].isna()) & (df["RANKING"] == 1)]
    print(f"  전체 REF: {rank1['REF_STYLE'].nunique()}개")
    print(f"  정답지 보유: {has_answer_rank1['REF_STYLE'].nunique()}개 (Wear only)")
    print(f"  정답지 없음: {no_answer_rank1['REF_STYLE'].nunique()}개 "
          f"(Wear {len(no_answer_rank1[no_answer_rank1['분류']=='Wear'])} "
          f"+ ACC {len(no_answer_rank1[no_answer_rank1['분류']=='ACC'])})")

    print("\n[2/13] 성능 요약 (Sheet 1)...")
    perf_df = build_performance_summary(df)

    print("[3/13] Miss 케이스 상세 (Sheet 2)...")
    miss_df = build_miss_analysis(df)
    print(f"  Miss 케이스: {len(miss_df)}건")

    print("[4/13] 정답지 없는 스타일 분석 (Sheet 3)...")
    no_answer_df = build_no_answer_analysis(df)
    print(f"  분석 대상: {len(no_answer_df)}건")

    print("[5/13] Wear vs ACC 비교 (Sheet 4)...")
    cat_comp_df = build_category_comparison(df, no_answer_df)

    print("[6/13] 최적화 방향 제안 (Sheet 5)...")
    opt_df = build_optimization_suggestions(df, miss_df, perf_df)

    print("[7/13] 하드필터 위반 검출 (Sheet 6)...")
    violations_df = build_hardfilter_violations(df)
    print(f"  위반 건수: {len(violations_df)}건")

    print("[8/13] Rank 변별력 분석 (Sheet 7)...")
    rank_summary_df, rank_detail_df = build_rank_discrimination(df)

    print("[9/13] 속성별 일치율 통계 (Sheet 8)...")
    attr_stats_df = build_attribute_stats(df)

    print("[10/13] 서브시즌 크로스매칭 (Sheet 9)...")
    subseason_cross_df = build_subseason_cross(df)

    print("[11/13] 서브시즌 시뮬레이션 (Sheet 10)...")
    sim_summary_df, sim_item_df, sim_improve_df = build_subseason_simulation(df)
    print(f"  Hit@1 신규 획득: {len(sim_improve_df)}건")

    print("[12/13] DOMAIN2 분석 (Sheet 11)...")
    domain2_df = build_domain2_analysis(df)

    print("[13/13] 용품 상세분석 (Sheet 12)...")
    acc_detail_df = build_acc_detail(df)

    # ── Excel 작성 ──
    print("\nExcel 보고서 작성 중...")
    wb = Workbook()

    # Sheet 1: 성능_요약
    ws1 = wb.active
    ws1.title = "성능_요약"
    write_df(ws1, perf_df)
    style_header(ws1, len(perf_df.columns))
    apply_borders(ws1)
    # 등급 색상
    grade_col = list(perf_df.columns).index("등급") + 1
    for r in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=r, column=grade_col)
        if cell.value == "A":
            cell.fill = GREEN_FILL
        elif cell.value == "B":
            cell.fill = YELLOW_FILL
        elif cell.value == "C":
            cell.fill = ORANGE_FILL
        elif cell.value == "D":
            cell.fill = RED_FILL
    # 전체 행 강조
    for c in range(1, len(perf_df.columns) + 1):
        ws1.cell(row=2, column=c).fill = SECTION_FILL
        ws1.cell(row=2, column=c).font = Font(bold=True, size=10)
    auto_width(ws1)
    ws1.freeze_panes = "A2"

    # Sheet 2: Miss_케이스_상세
    ws2 = wb.create_sheet("Miss_케이스_상세")
    if len(miss_df) > 0:
        write_df(ws2, miss_df)
        style_header(ws2, len(miss_df.columns))
        apply_borders(ws2)
        # Miss 유형 색상
        mt_col = list(miss_df.columns).index("Miss_유형") + 1
        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=mt_col)
            if cell.value == "유사실패":
                cell.fill = RED_FILL
            elif cell.value and "순위밀림" in str(cell.value):
                cell.fill = YELLOW_FILL
        # DOMAIN 일치 색상
        dom_col = list(miss_df.columns).index("REF↔Rank1_DOMAIN일치") + 1
        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=dom_col)
            if cell.value == "O":
                cell.fill = GREEN_FILL
            elif cell.value == "X":
                cell.fill = RED_FILL
    else:
        ws2.cell(row=1, column=1, value="Miss 케이스 없음")
    auto_width(ws2)
    ws2.freeze_panes = "A2"

    # Sheet 3: 정답지_없는_스타일_분석
    ws3 = wb.create_sheet("정답지_없는_스타일_분석")
    if len(no_answer_df) > 0:
        write_df(ws3, no_answer_df)
        style_header(ws3, len(no_answer_df.columns))
        apply_borders(ws3)
        # 위험플래그 색상
        risk_col = list(no_answer_df.columns).index("위험플래그") + 1
        for r in range(2, ws3.max_row + 1):
            cell = ws3.cell(row=r, column=risk_col)
            if cell.value == "양호":
                cell.fill = GREEN_FILL
            elif cell.value == "점검필요":
                cell.fill = YELLOW_FILL
            elif cell.value == "불량추정":
                cell.fill = RED_FILL
        # 일관성 O/X 색상
        for col_name in ["DOMAIN일관", "FIT일관", "가격일관"]:
            if col_name in no_answer_df.columns:
                ci = list(no_answer_df.columns).index(col_name) + 1
                for r in range(2, ws3.max_row + 1):
                    cell = ws3.cell(row=r, column=ci)
                    if cell.value == "O":
                        cell.fill = GREEN_FILL
                    elif cell.value == "X":
                        cell.fill = RED_FILL
    auto_width(ws3)
    ws3.freeze_panes = "A2"

    # Sheet 4: Wear_vs_ACC_비교
    ws4 = wb.create_sheet("Wear_vs_ACC_비교")
    write_df(ws4, cat_comp_df)
    style_header(ws4, len(cat_comp_df.columns))
    apply_borders(ws4)
    # 섹션 헤더 스타일
    for r in range(2, ws4.max_row + 1):
        cell = ws4.cell(row=r, column=1)
        if cell.value and str(cell.value).startswith("──"):
            for c in range(1, len(cat_comp_df.columns) + 1):
                ws4.cell(row=r, column=c).fill = SECTION_FILL
                ws4.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws4)

    # Sheet 5: 최적화_방향_제안
    ws5 = wb.create_sheet("최적화_방향_제안")
    write_df(ws5, opt_df)
    style_header(ws5, len(opt_df.columns))
    apply_borders(ws5)
    # 주제 섹션 헤더 스타일
    for r in range(2, ws5.max_row + 1):
        cell = ws5.cell(row=r, column=1)
        val = str(cell.value) if cell.value else ""
        if val and val[0].isdigit() and "." in val:
            for c in range(1, len(opt_df.columns) + 1):
                ws5.cell(row=r, column=c).fill = SECTION_FILL
                ws5.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws5, max_width=60)

    # Sheet 6: 하드필터_위반_검출
    ws6 = wb.create_sheet("하드필터_위반_검출")
    if len(violations_df) > 0:
        write_df(ws6, violations_df)
        style_header(ws6, len(violations_df.columns))
        apply_borders(ws6)
        sev_col = list(violations_df.columns).index("심각도") + 1
        for r in range(2, ws6.max_row + 1):
            cell = ws6.cell(row=r, column=sev_col)
            if cell.value == "치명":
                cell.fill = RED_FILL
            elif cell.value == "주의":
                cell.fill = YELLOW_FILL
    else:
        ws6.cell(row=1, column=1, value="하드필터 위반 케이스 없음")
    auto_width(ws6)
    ws6.freeze_panes = "A2"

    # Sheet 7: Rank_변별력_분석
    ws7 = wb.create_sheet("Rank_변별력_분석")
    # 요약
    ws7.cell(row=1, column=1, value="Rank별 평균 지표").font = Font(bold=True, size=11)
    ws7.cell(row=1, column=1).fill = SECTION_FILL
    ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    write_df(ws7, rank_summary_df, start_row=2)
    style_header(ws7, len(rank_summary_df.columns), row=2)
    for r in range(3, 3 + len(rank_summary_df)):
        for c in range(1, len(rank_summary_df.columns) + 1):
            ws7.cell(row=r, column=c).border = THIN_BORDER
            ws7.cell(row=r, column=c).font = CELL_FONT
    # 상세
    detail_start = 3 + len(rank_summary_df) + 2
    ws7.cell(row=detail_start, column=1, value="REF별 Rank 역전 상세").font = Font(bold=True, size=11)
    ws7.cell(row=detail_start, column=1).fill = SECTION_FILL
    ws7.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=5)
    write_df(ws7, rank_detail_df, start_row=detail_start + 1)
    style_header(ws7, len(rank_detail_df.columns), row=detail_start + 1)
    for r in range(detail_start + 2, detail_start + 2 + len(rank_detail_df)):
        for c in range(1, len(rank_detail_df.columns) + 1):
            ws7.cell(row=r, column=c).border = THIN_BORDER
            ws7.cell(row=r, column=c).font = CELL_FONT
        # 역전 색상
        inv_col = len(rank_detail_df.columns)
        cell = ws7.cell(row=r, column=inv_col)
        if cell.value == "Rank3>Rank1":
            cell.fill = RED_FILL
        elif cell.value == "차이없음":
            cell.fill = YELLOW_FILL
    # O/X 색상
    for r in range(detail_start + 2, detail_start + 2 + len(rank_detail_df)):
        for c in range(1, len(rank_detail_df.columns) + 1):
            cell = ws7.cell(row=r, column=c)
            if cell.value == "O":
                cell.fill = GREEN_FILL
            elif cell.value == "X":
                cell.fill = RED_FILL
    auto_width(ws7)

    # Sheet 8: 속성별_일치율_통계
    ws8 = wb.create_sheet("속성별_일치율_통계")
    write_df(ws8, attr_stats_df)
    style_header(ws8, len(attr_stats_df.columns))
    apply_borders(ws8)
    for r in range(2, ws8.max_row + 1):
        cell = ws8.cell(row=r, column=1)
        if cell.value and str(cell.value).startswith("──"):
            for c in range(1, len(attr_stats_df.columns) + 1):
                ws8.cell(row=r, column=c).fill = SECTION_FILL
                ws8.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws8)

    # Sheet 9: 서브시즌_크로스매칭
    ws9 = wb.create_sheet("서브시즌_크로스매칭")
    if len(subseason_cross_df) > 0:
        write_df(ws9, subseason_cross_df)
        style_header(ws9, len(subseason_cross_df.columns))
        apply_borders(ws9)
        for r in range(2, ws9.max_row + 1):
            cell = ws9.cell(row=r, column=1)
            if cell.value and str(cell.value).startswith("──"):
                for c in range(1, len(subseason_cross_df.columns) + 1):
                    ws9.cell(row=r, column=c).fill = SECTION_FILL
                    ws9.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws9)

    # Sheet 10: 서브시즌_시뮬레이션
    ws10 = wb.create_sheet("서브시즌_시뮬레이션")
    # 요약 테이블
    ws10.cell(row=1, column=1, value="AS-IS vs TO-BE 성능 비교").font = Font(bold=True, size=11)
    ws10.cell(row=1, column=1).fill = SECTION_FILL
    ws10.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    write_df(ws10, sim_summary_df, start_row=2)
    style_header(ws10, len(sim_summary_df.columns), row=2)
    for r in range(3, 3 + len(sim_summary_df)):
        for c in range(1, len(sim_summary_df.columns) + 1):
            ws10.cell(row=r, column=c).border = THIN_BORDER
            ws10.cell(row=r, column=c).font = CELL_FONT
    # 개선효과 행 강조
    effect_row = 3 + len(sim_summary_df) - 1
    for c in range(1, len(sim_summary_df.columns) + 1):
        ws10.cell(row=effect_row, column=c).fill = GREEN_FILL
        ws10.cell(row=effect_row, column=c).font = Font(bold=True, size=10)
    # ITEM별 시뮬레이션
    item_start = 3 + len(sim_summary_df) + 2
    ws10.cell(row=item_start, column=1, value="ITEM별 시뮬레이션").font = Font(bold=True, size=11)
    ws10.cell(row=item_start, column=1).fill = SECTION_FILL
    ws10.merge_cells(start_row=item_start, start_column=1, end_row=item_start, end_column=5)
    if len(sim_item_df) > 0:
        write_df(ws10, sim_item_df, start_row=item_start + 1)
        style_header(ws10, len(sim_item_df.columns), row=item_start + 1)
        for r in range(item_start + 2, item_start + 2 + len(sim_item_df)):
            for c in range(1, len(sim_item_df.columns) + 1):
                ws10.cell(row=r, column=c).border = THIN_BORDER
                ws10.cell(row=r, column=c).font = CELL_FONT
            # delta > 0 강조
            delta_col = list(sim_item_df.columns).index("delta") + 1
            cell = ws10.cell(row=r, column=delta_col)
            if cell.value and str(cell.value).startswith("+") and cell.value != "+0":
                cell.fill = GREEN_FILL
            # 등급변화에 화살표 있으면 강조
            gc_col = list(sim_item_df.columns).index("등급변화") + 1
            gc_cell = ws10.cell(row=r, column=gc_col)
            if gc_cell.value and "→" in str(gc_cell.value):
                gc_cell.fill = YELLOW_FILL
    # Hit@1 신규 획득 케이스
    if len(sim_improve_df) > 0:
        imp_start = item_start + 2 + len(sim_item_df) + 2
        ws10.cell(row=imp_start, column=1, value="Hit@1 신규 획득 케이스").font = Font(bold=True, size=11)
        ws10.cell(row=imp_start, column=1).fill = SECTION_FILL
        ws10.merge_cells(start_row=imp_start, start_column=1, end_row=imp_start, end_column=5)
        write_df(ws10, sim_improve_df, start_row=imp_start + 1)
        style_header(ws10, len(sim_improve_df.columns), row=imp_start + 1)
        for r in range(imp_start + 2, imp_start + 2 + len(sim_improve_df)):
            for c in range(1, len(sim_improve_df.columns) + 1):
                ws10.cell(row=r, column=c).border = THIN_BORDER
                ws10.cell(row=r, column=c).font = CELL_FONT
    auto_width(ws10)

    # Sheet 11: DOMAIN2_분석
    ws11 = wb.create_sheet("DOMAIN2_분석")
    if len(domain2_df) > 0:
        write_df(ws11, domain2_df)
        style_header(ws11, len(domain2_df.columns))
        apply_borders(ws11)
        for r in range(2, ws11.max_row + 1):
            cell = ws11.cell(row=r, column=1)
            if cell.value and str(cell.value).startswith("──"):
                for c in range(1, len(domain2_df.columns) + 1):
                    ws11.cell(row=r, column=c).fill = SECTION_FILL
                    ws11.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws11, max_width=60)

    # Sheet 12: 용품_상세분석
    ws12 = wb.create_sheet("용품_상세분석")
    if len(acc_detail_df) > 0:
        write_df(ws12, acc_detail_df)
        style_header(ws12, len(acc_detail_df.columns))
        apply_borders(ws12)
        for r in range(2, ws12.max_row + 1):
            cell = ws12.cell(row=r, column=1)
            if cell.value and str(cell.value).startswith("──"):
                for c in range(1, len(acc_detail_df.columns) + 1):
                    ws12.cell(row=r, column=c).fill = SECTION_FILL
                    ws12.cell(row=r, column=c).font = Font(bold=True, size=10)
    auto_width(ws12, max_width=60)

    wb.save(OUTPUT_PATH)
    print(f"\n→ 저장 완료: {OUTPUT_PATH}")

    # ── 콘솔 요약 ──
    print("\n" + "=" * 60)
    print("성능 측정 결과 요약")
    print("=" * 60)

    print("\n[정답지 기반 성능]")
    for _, r in perf_df.iterrows():
        if r["구분"] == "전체":
            print(f"  전체: Hit@1={r['Hit@1']}, Hit@3={r['Hit@3']}, MRR={r['MRR']}")
            break

    print("\n[ITEM별 성능 (Hit@1 기준)]")
    for _, r in perf_df.iterrows():
        if r["구분"] == "전체":
            continue
        grade = r.get("등급", "")
        print(f"  {r['구분']:20s} Hit@1={r['Hit@1']:15s} Miss={r['Miss']:15s} 등급={grade}")

    if len(miss_df) > 0:
        print(f"\n[Miss 분석] 총 {len(miss_df)}건")
        miss_types = miss_df["Miss_유형"].value_counts()
        for t, c in miss_types.items():
            print(f"  {t}: {c}건")

    if len(no_answer_df) > 0:
        print(f"\n[정답지 없는 REF 프록시 분석] 총 {len(no_answer_df)}건")
        risk_dist = no_answer_df["위험플래그"].value_counts()
        for r, c in risk_dist.items():
            print(f"  {r}: {c}건 ({c/len(no_answer_df)*100:.1f}%)")

    print(f"\n[하드필터 위반] 총 {len(violations_df)}건")
    if len(violations_df) > 0:
        sev = violations_df["심각도"].value_counts()
        if "치명" in sev:
            print(f"  치명(성별 위반): {sev['치명']}건")
        if "주의" in sev:
            print(f"  주의(가격 극단 등): {sev['주의']}건")

    print(f"\n[Rank 변별력]")
    for _, r in rank_summary_df.iterrows():
        val = r.iloc[0] if isinstance(r.iloc[0], str) else str(r.iloc[0])
        if val.startswith("Rank "):
            print(f"  {val}: DOMAIN {r['DOMAIN 일치율']}, "
                  f"상품명유사도 {r['평균 상품명유사도']}, "
                  f"가격차이 {r['평균 가격차이%']}")
    if len(rank_detail_df) > 0 and "Rank역전" in rank_detail_df.columns:
        inv = (rank_detail_df["Rank역전"] == "Rank3>Rank1").sum()
        nodiff = (rank_detail_df["Rank역전"] == "차이없음").sum()
        total_ref = len(rank_detail_df)
        print(f"  역전(3>1): {inv}/{total_ref} ({inv/total_ref*100:.1f}%)")
        print(f"  차이없음: {nodiff}/{total_ref} ({nodiff/total_ref*100:.1f}%)")

    # 서브시즌 시뮬레이션 핵심 수치
    if len(sim_summary_df) >= 3:
        print(f"\n[서브시즌 하드필터 시뮬레이션]")
        asis_row = sim_summary_df.iloc[0]
        tobe_row = sim_summary_df.iloc[1]
        effect_row_data = sim_summary_df.iloc[2]
        print(f"  AS-IS: Hit@1={asis_row['Hit@1']}, MRR={asis_row['MRR']}")
        print(f"  TO-BE: Hit@1={tobe_row['Hit@1']}, MRR={tobe_row['MRR']}")
        print(f"  개선효과: Hit@1 {effect_row_data['Hit@1']}, MRR {effect_row_data['MRR']}")
        if len(sim_improve_df) > 0:
            print(f"  Hit@1 신규 획득: {len(sim_improve_df)}건")

    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
