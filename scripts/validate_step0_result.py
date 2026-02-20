"""
STEP 0 유사스타일 1차 모델 정확도 검증 (현업 리뷰 중심)

목적: REF 스타일과 SIMILAR 스타일이 실제로 유사한 상품인지 현업이 판단할 수 있는
      리뷰 시트를 만들고, 모델의 체계적 실패 패턴과 변별력을 분석한다.

- input:  output/result_1st_draft.xlsx
- output: output/step0_validation_report.xlsx (5 sheets)
"""

import os
import re
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── paths ──────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_PATH = os.path.join(BASE_DIR, "output", "result_1st_draft.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "output", "step0_validation_report.xlsx")


# ── helpers ────────────────────────────────────────────────────


def _clean(v):
    """None / 0 / 'None' / NaN → None, else stripped string."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v).strip()
    if s in ("", "0", "None", "nan", "NaN"):
        return None
    return s


def text_similarity(a, b):
    a, b = _clean(a), _clean(b)
    if a is None or b is None:
        return None
    a = re.sub(r"[^\w\s]", "", a)
    b = re.sub(r"[^\w\s]", "", b)
    return round(SequenceMatcher(None, a, b).ratio(), 3)


def price_diff_pct(ref_price, sim_price):
    """가격 차이 % (0=동일, 100=한쪽이 2배)."""
    rp, sp = _clean(ref_price), _clean(sim_price)
    if rp is None or sp is None:
        return None
    try:
        rp, sp = float(rp), float(sp)
    except ValueError:
        return None
    denom = max(rp, sp)
    if denom == 0:
        return 0.0
    return round(abs(rp - sp) / denom * 100, 1)


def gender_compatible(ref_g, sim_g):
    rg, sg = _clean(ref_g), _clean(sim_g)
    if rg is None or sg is None:
        return None
    if rg == sg:
        return True
    if "공용" in (rg, sg):
        return True
    return False


# ── data loading ───────────────────────────────────────────────


def load_data():
    df = pd.read_excel(INPUT_PATH, sheet_name="result_1st_draft", engine="openpyxl")
    unnamed = [c for c in df.columns if "Unnamed" in str(c)]
    if unnamed:
        df = df.drop(columns=unnamed)
    df = df.rename(columns={
        "PRDT_NM": "REF_PRDT_NM",
        "PRDT_NM.1": "SIM_PRDT_NM",
        "성별": "REF_SEX_NM",
        "TAG": "REF_TAG_PRICE",
        "FIT": "REF_FIT",
        "DOMAIN": "REF_DOMAIN",
        "FAB": "REF_FAB",
    })
    return df


# ── Sheet 1: 현업 리뷰 시트 ───────────────────────────────────


def build_review_sheet(df):
    """
    현업이 직접 "이 매칭 맞아/틀려"를 판정할 수 있는 시트.
    REF 스타일별로 Rank 1~3을 묶어서 보여주고,
    판정(O/X)과 틀린이유 컬럼을 비워둔다.
    """
    rows = []
    for ref_style, group in df.groupby("REF_STYLE_CD", sort=False):
        group = group.sort_values("RANKING")
        ref_row = group.iloc[0]
        for _, row in group.iterrows():
            pdiff = price_diff_pct(row["REF_TAG_PRICE"], row["TAG_PRICE"])
            rows.append({
                "분류": row["분류"],
                "REF 스타일": row["REF_STYLE_CD"],
                "REF 상품명": row["REF_PRDT_NM"],
                "REF 성별": row["REF_SEX_NM"],
                "REF 가격": row["REF_TAG_PRICE"],
                "REF FIT": _clean(row["REF_FIT"]) or "",
                "REF DOMAIN": _clean(row["REF_DOMAIN"]) or "",
                "": "",  # separator
                "Rank": int(row["RANKING"]),
                "SIM 스타일": row["SIMILAR_STYLE_CD"],
                "SIM 상품명": row["SIM_PRDT_NM"],
                "SIM 성별": row["SEX_NM"],
                "SIM 가격": row["TAG_PRICE"],
                "SIM FIT": _clean(row["FIT_INFO1"]) or "",
                "SIM DOMAIN": _clean(row["DOMAIN1_NM"]) or "",
                "가격차이%": f"{pdiff:.0f}%" if pdiff is not None else "",
                " ": "",  # separator
                "판정": "",       # ← 현업이 O/X 입력
                "틀린이유": "",   # ← 현업이 사유 기입
            })

    return pd.DataFrame(rows)


# ── Sheet 2: 하드필터 위반 검출 ────────────────────────────────


def build_hardfilter_violations(df):
    """
    모델이 지켜야 할 기본 규칙을 위반한 케이스.
    위반 = 모델 버그. 이건 가중치 문제가 아니라 로직 수정이 필요.
    """
    violations = []

    for _, row in df.iterrows():
        issues = []

        # 1) 성별 위반: 남성↔여성 (공용 없이)
        if not gender_compatible(row["REF_SEX_NM"], row["SEX_NM"]):
            issues.append(f"성별 위반: {_clean(row['REF_SEX_NM'])}→{_clean(row['SEX_NM'])}")

        # 2) 극단적 가격 차이 (50% 이상) → 하드필터는 아니지만 의심 케이스
        pdiff = price_diff_pct(row["REF_TAG_PRICE"], row["TAG_PRICE"])
        if pdiff is not None and pdiff >= 50:
            issues.append(f"가격 차이 극단: {pdiff:.0f}% ({row['REF_TAG_PRICE']:,}→{row['TAG_PRICE']:,})")

        # 3) DOMAIN 교차 매칭 중 비즈니스적으로 문제되는 패턴
        ref_d = _clean(row["REF_DOMAIN"])
        sim_d = _clean(row["DOMAIN1_NM"])
        if ref_d and sim_d and ref_d != sim_d:
            # Monogram↔Basic은 디자인 성격이 완전히 다름
            pair = frozenset([ref_d, sim_d])
            problematic_pairs = [
                frozenset(["Monogram", "Varsity"]),
                frozenset(["Megagram", "Basic"]),
            ]
            if pair in problematic_pairs:
                issues.append(f"DOMAIN 교차: {ref_d}→{sim_d}")

        if issues:
            violations.append({
                "Rank": int(row["RANKING"]),
                "분류": row["분류"],
                "REF 스타일": row["REF_STYLE_CD"],
                "REF 상품명": row["REF_PRDT_NM"],
                "REF 성별": _clean(row["REF_SEX_NM"]),
                "REF 가격": row["REF_TAG_PRICE"],
                "REF DOMAIN": _clean(row["REF_DOMAIN"]),
                "SIM 스타일": row["SIMILAR_STYLE_CD"],
                "SIM 상품명": row["SIM_PRDT_NM"],
                "SIM 성별": _clean(row["SEX_NM"]),
                "SIM 가격": row["TAG_PRICE"],
                "SIM DOMAIN": _clean(row["DOMAIN1_NM"]),
                "위반 내용": " | ".join(issues),
                "심각도": "치명" if any("성별" in i for i in issues) else "주의",
            })

    result = pd.DataFrame(violations)
    if len(result) > 0:
        # 치명 먼저, 그 다음 주의
        result = result.sort_values(
            ["심각도", "분류", "REF 스타일"],
            ascending=[True, True, True],
        ).reset_index(drop=True)
    return result


# ── Sheet 3: Rank 변별력 분석 ──────────────────────────────────


def build_rank_discrimination(df):
    """
    모델의 Rank 1이 정말 Rank 3보다 나은지.
    Rank 간 차이가 없으면 = 스코어링 함수의 변별력이 부족한 것.
    """
    # REF 스타일별로 Rank 1/2/3의 속성 비교
    ref_groups = df.groupby("REF_STYLE_CD")

    per_ref_rows = []
    for ref_style, group in ref_groups:
        group = group.sort_values("RANKING")
        ref_row = group.iloc[0]

        rank_data = {"REF 스타일": ref_style, "REF 상품명": ref_row["REF_PRDT_NM"],
                     "분류": ref_row["분류"]}

        sims = {}
        for _, row in group.iterrows():
            rank = int(row["RANKING"])
            # 이 Rank의 SIM이 REF와 얼마나 비슷한지 간단 지표들
            domain_ok = 1 if _clean(row["REF_DOMAIN"]) == _clean(row["DOMAIN1_NM"]) else 0
            name_sim = text_similarity(row["REF_PRDT_NM"], row["SIM_PRDT_NM"])
            pdiff = price_diff_pct(row["REF_TAG_PRICE"], row["TAG_PRICE"])

            sims[rank] = {
                "style": row["SIMILAR_STYLE_CD"],
                "name": row["SIM_PRDT_NM"],
                "domain_ok": domain_ok,
                "name_sim": name_sim or 0,
                "price_diff": pdiff or 0,
            }

        # Rank 1~3 정보
        for r in [1, 2, 3]:
            if r in sims:
                rank_data[f"Rank{r} 스타일"] = sims[r]["style"]
                rank_data[f"Rank{r} 상품명"] = sims[r]["name"]
                rank_data[f"Rank{r} DOMAIN일치"] = "O" if sims[r]["domain_ok"] else "X"
                rank_data[f"Rank{r} 상품명유사도"] = sims[r]["name_sim"]
                rank_data[f"Rank{r} 가격차이%"] = sims[r]["price_diff"]

        # Rank 1 vs 3 역전 여부 (Rank 3가 더 나은 케이스)
        if 1 in sims and 3 in sims:
            # Rank 3의 상품명 유사도가 더 높고 + 가격도 더 가까우면 = 역전
            r1, r3 = sims[1], sims[3]
            r1_better = (r1["name_sim"] > r3["name_sim"]) or (r1["price_diff"] < r3["price_diff"])
            r3_better = (r3["name_sim"] > r1["name_sim"] + 0.1) and (r3["price_diff"] < r1["price_diff"])
            if r3_better:
                rank_data["Rank역전"] = "Rank3>Rank1"
            elif not r1_better:
                rank_data["Rank역전"] = "차이없음"
            else:
                rank_data["Rank역전"] = ""
        else:
            rank_data["Rank역전"] = ""

        per_ref_rows.append(rank_data)

    per_ref_df = pd.DataFrame(per_ref_rows)

    # 요약 통계
    summary_rows = []

    # 전체 Rank별 평균
    for rank in [1, 2, 3]:
        rank_df = df[df["RANKING"] == rank]
        domain_rate = sum(
            1 for _, r in rank_df.iterrows()
            if _clean(r["REF_DOMAIN"]) == _clean(r["DOMAIN1_NM"])
        ) / len(rank_df) * 100

        name_sims = [
            text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
            for _, r in rank_df.iterrows()
        ]
        name_sims = [x for x in name_sims if x is not None]
        avg_name = sum(name_sims) / len(name_sims) if name_sims else 0

        pdiffs = [
            price_diff_pct(r["REF_TAG_PRICE"], r["TAG_PRICE"])
            for _, r in rank_df.iterrows()
        ]
        pdiffs = [x for x in pdiffs if x is not None]
        avg_pdiff = sum(pdiffs) / len(pdiffs) if pdiffs else 0

        summary_rows.append({
            "": f"Rank {rank}",
            "건수": len(rank_df),
            "DOMAIN 일치율": f"{domain_rate:.1f}%",
            "평균 상품명유사도": f"{avg_name:.3f}",
            "평균 가격차이%": f"{avg_pdiff:.1f}%",
        })

    # Rank 1→3 변화
    if len(summary_rows) >= 3:
        summary_rows.append({
            "": "Rank1→3 변화",
            "건수": "",
            "DOMAIN 일치율": "",
            "평균 상품명유사도": f"Rank1과 3의 차이가 작으면 = 변별력 부족",
            "평균 가격차이%": "",
        })

    # 역전 건수
    if len(per_ref_df) > 0 and "Rank역전" in per_ref_df.columns:
        inversions = (per_ref_df["Rank역전"] == "Rank3>Rank1").sum()
        no_diff = (per_ref_df["Rank역전"] == "차이없음").sum()
        total = len(per_ref_df)
        summary_rows.append({
            "": "Rank 역전 (3>1)",
            "건수": f"{inversions}건 / {total}개 REF ({inversions/total*100:.1f}%)",
            "DOMAIN 일치율": "",
            "평균 상품명유사도": "",
            "평균 가격차이%": "",
        })
        summary_rows.append({
            "": "Rank 차이없음",
            "건수": f"{no_diff}건 / {total}개 REF ({no_diff/total*100:.1f}%)",
            "DOMAIN 일치율": "",
            "평균 상품명유사도": "",
            "평균 가격차이%": "",
        })

    summary_df = pd.DataFrame(summary_rows)
    return summary_df, per_ref_df


# ── Sheet 4: 체계적 실패 패턴 ──────────────────────────────────


def build_failure_patterns(df):
    """
    모델이 체계적으로 실패하는 패턴을 식별.
    "어떤 유형의 REF에서 모델이 엉뚱한 SIM을 반환하는가?"
    """
    rows = []

    # 1) DOMAIN 불일치 패턴 (빈도순)
    domain_mismatches = []
    for _, row in df.iterrows():
        ref_d = _clean(row["REF_DOMAIN"])
        sim_d = _clean(row["DOMAIN1_NM"])
        if ref_d and sim_d and ref_d != sim_d:
            domain_mismatches.append((ref_d, sim_d))

    dm_counts = pd.Series(domain_mismatches).value_counts()
    domain_rows = []
    for (ref_d, sim_d), cnt in dm_counts.items():
        domain_rows.append({
            "패턴 유형": "DOMAIN 불일치",
            "REF 값": ref_d,
            "SIM 값": sim_d,
            "건수": cnt,
            "해석": (
                "디자인 계열이 다른 상품 매칭 → DOMAIN 가중치 상향 검토"
                if cnt >= 10 else ""
            ),
        })

    # 2) 상품명 유사도 극히 낮은 케이스 패턴
    low_name_rows = []
    for _, row in df.iterrows():
        ns = text_similarity(row["REF_PRDT_NM"], row["SIM_PRDT_NM"])
        if ns is not None and ns < 0.2:
            low_name_rows.append({
                "패턴 유형": "상품명 유사도 < 0.2",
                "REF 값": row["REF_PRDT_NM"],
                "SIM 값": row["SIM_PRDT_NM"],
                "건수": 1,
                "해석": f"유사도={ns:.2f}, 분류={row['분류']}",
            })

    # 3) 같은 SIM이 여러 다른 REF에 반복 매칭되는 패턴 (모델이 특정 상품을 과도 추천)
    sim_freq = df["SIMILAR_STYLE_CD"].value_counts()
    overused = sim_freq[sim_freq >= 10]
    overuse_rows = []
    for sim_style, cnt in overused.items():
        matched_refs = df[df["SIMILAR_STYLE_CD"] == sim_style]
        sim_name = matched_refs.iloc[0]["SIM_PRDT_NM"]
        ref_count = matched_refs["REF_STYLE_CD"].nunique()
        overuse_rows.append({
            "패턴 유형": "SIM 과다 추천",
            "REF 값": f"{ref_count}개 REF에 매칭",
            "SIM 값": f"{sim_style} ({sim_name})",
            "건수": cnt,
            "해석": "특정 상품이 너무 많은 REF에 반환됨 → 다양성 부족",
        })
    overuse_rows.sort(key=lambda x: x["건수"], reverse=True)

    # 4) Rank 1인데 DOMAIN 불일치 (가장 유사하다고 한 건데 기본 속성이 다름)
    rank1_domain_fail = []
    rank1 = df[df["RANKING"] == 1]
    for _, row in rank1.iterrows():
        ref_d = _clean(row["REF_DOMAIN"])
        sim_d = _clean(row["DOMAIN1_NM"])
        if ref_d and sim_d and ref_d != sim_d:
            rank1_domain_fail.append({
                "패턴 유형": "Rank1인데 DOMAIN 불일치",
                "REF 값": f"{row['REF_STYLE_CD']} ({row['REF_PRDT_NM']}, {ref_d})",
                "SIM 값": f"{row['SIMILAR_STYLE_CD']} ({row['SIM_PRDT_NM']}, {sim_d})",
                "건수": 1,
                "해석": "최우선 추천인데 디자인 계열이 다름 → 우선 개선 대상",
            })

    # 합치기
    all_patterns = (
        domain_rows
        + [{"패턴 유형": "───", "REF 값": "", "SIM 값": "", "건수": "", "해석": ""}]
        + overuse_rows
        + [{"패턴 유형": "───", "REF 값": "", "SIM 값": "", "건수": "", "해석": ""}]
        + rank1_domain_fail[:20]  # 상위 20건만
        + [{"패턴 유형": "───", "REF 값": "", "SIM 값": "", "건수": "", "해석": ""}]
        + low_name_rows[:20]  # 상위 20건만
    )

    return pd.DataFrame(all_patterns)


# ── Sheet 5: 속성별 참고 통계 ──────────────────────────────────


def build_reference_stats(df):
    """속성별 기초 통계 (참고용)."""
    total = len(df)

    # DOMAIN
    domain_match = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_DOMAIN"]) == _clean(r["DOMAIN1_NM"])
        and _clean(r["REF_DOMAIN"]) is not None
    )
    domain_total = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_DOMAIN"]) is not None and _clean(r["DOMAIN1_NM"]) is not None
    )

    # FIT
    fit_exact = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_FIT"]) is not None
        and _clean(r["FIT_INFO1"]) is not None
        and _clean(r["REF_FIT"]) == _clean(r["FIT_INFO1"])
    )
    fit_total = sum(
        1 for _, r in df.iterrows()
        if _clean(r["REF_FIT"]) is not None and _clean(r["FIT_INFO1"]) is not None
    )

    # Gender
    gender_ok = sum(
        1 for _, r in df.iterrows()
        if gender_compatible(r["REF_SEX_NM"], r["SEX_NM"]) is True
    )

    # Price
    pdiffs = [
        price_diff_pct(r["REF_TAG_PRICE"], r["TAG_PRICE"])
        for _, r in df.iterrows()
    ]
    pdiffs = [x for x in pdiffs if x is not None]

    # Name similarity
    nsims = [
        text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
        for _, r in df.iterrows()
    ]
    nsims = [x for x in nsims if x is not None]

    stats = [
        {"속성": "DOMAIN 일치율", "전체": f"{domain_match/domain_total*100:.1f}%",
         "건수": f"{domain_match}/{domain_total}",
         "비고": "Exact match 기준"},
        {"속성": "FIT 완전일치율", "전체": f"{fit_exact/fit_total*100:.1f}%",
         "건수": f"{fit_exact}/{fit_total}",
         "비고": "값 존재하는 건만 대상"},
        {"속성": "성별 호환율", "전체": f"{gender_ok/total*100:.1f}%",
         "건수": f"{gender_ok}/{total}",
         "비고": "공용↔남녀 허용"},
        {"속성": "가격 평균차이", "전체": f"{sum(pdiffs)/len(pdiffs):.1f}%",
         "건수": f"{len(pdiffs)}건",
         "비고": "max(REF,SIM) 대비 차이율"},
        {"속성": "상품명 평균유사도", "전체": f"{sum(nsims)/len(nsims):.3f}",
         "건수": f"{len(nsims)}건",
         "비고": "SequenceMatcher 기준"},
    ]

    # Wear vs ACC breakdown
    for cat in ["Wear", "ACC"]:
        cat_df = df[df["분류"] == cat]
        if len(cat_df) == 0:
            continue
        cat_nsims = [
            text_similarity(r["REF_PRDT_NM"], r["SIM_PRDT_NM"])
            for _, r in cat_df.iterrows()
        ]
        cat_nsims = [x for x in cat_nsims if x is not None]
        cat_pdiffs = [
            price_diff_pct(r["REF_TAG_PRICE"], r["TAG_PRICE"])
            for _, r in cat_df.iterrows()
        ]
        cat_pdiffs = [x for x in cat_pdiffs if x is not None]
        stats.append({
            "속성": f"── {cat} ──",
            "전체": "",
            "건수": f"{len(cat_df)}건",
            "비고": "",
        })
        stats.append({
            "속성": f"  상품명 평균유사도",
            "전체": f"{sum(cat_nsims)/len(cat_nsims):.3f}" if cat_nsims else "-",
            "건수": "",
            "비고": "",
        })
        stats.append({
            "속성": f"  가격 평균차이",
            "전체": f"{sum(cat_pdiffs)/len(cat_pdiffs):.1f}%" if cat_pdiffs else "-",
            "건수": "",
            "비고": "",
        })

    return pd.DataFrame(stats)


# ── Excel formatting ───────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SEP_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
JUDGE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, min_width=8, max_width=35):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def apply_borders(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = CELL_FONT


def write_df(ws, df, start_row=1):
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, float) and pd.isna(val):
                ws.cell(row=r_idx, column=c_idx, value="")
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)


# ── main ───────────────────────────────────────────────────────


def main():
    print("=" * 60)
    print("STEP 0 유사스타일 1차 모델 정확도 검증 (현업 리뷰 중심)")
    print("=" * 60)

    print("\n[1/6] 데이터 로딩...")
    df = load_data()
    print(f"  총 {len(df)}건 (REF {df['REF_STYLE_CD'].nunique()}개 스타일, Rank 1~3)")

    print("[2/6] 현업 리뷰 시트...")
    review_df = build_review_sheet(df)

    print("[3/6] 하드필터 위반 검출...")
    violations_df = build_hardfilter_violations(df)

    print("[4/6] Rank 변별력 분석...")
    rank_summary_df, rank_detail_df = build_rank_discrimination(df)

    print("[5/6] 체계적 실패 패턴...")
    patterns_df = build_failure_patterns(df)

    print("[6/6] Excel 보고서 작성...")
    wb = Workbook()

    # ── Sheet 1: 현업 리뷰 시트
    ws1 = wb.active
    ws1.title = "현업_리뷰_시트"
    write_df(ws1, review_df)
    style_header(ws1, len(review_df.columns))
    apply_borders(ws1)

    # 판정/틀린이유 컬럼 강조
    judge_col = list(review_df.columns).index("판정") + 1
    reason_col = list(review_df.columns).index("틀린이유") + 1
    for r in range(1, ws1.max_row + 1):
        ws1.cell(row=r, column=judge_col).fill = JUDGE_FILL
        ws1.cell(row=r, column=reason_col).fill = JUDGE_FILL

    # 판정 컬럼에 드롭다운 (O/X/△)
    dv = DataValidation(type="list", formula1='"O,X,△"', allow_blank=True)
    dv.error = "O(유사), X(비유사), △(애매) 중 선택"
    dv.errorTitle = "판정"
    ws1.add_data_validation(dv)
    for r in range(2, ws1.max_row + 1):
        dv.add(ws1.cell(row=r, column=judge_col))

    # REF 스타일별 3행 그룹 구분 (짝수/홀수 REF에 배경색)
    ref_styles = review_df["REF 스타일"].unique()
    for i, ref_style in enumerate(ref_styles):
        if i % 2 == 1:
            # 이 REF의 행 범위
            mask = review_df["REF 스타일"] == ref_style
            row_indices = mask[mask].index
            for ri in row_indices:
                for c in range(1, 8):  # REF 컬럼들만 배경색
                    ws1.cell(row=ri + 2, column=c).fill = SEP_FILL

    auto_width(ws1)
    ws1.freeze_panes = "I2"  # REF 컬럼 고정
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: 하드필터 위반
    ws2 = wb.create_sheet("하드필터_위반_검출")
    if len(violations_df) > 0:
        write_df(ws2, violations_df)
        style_header(ws2, len(violations_df.columns))
        apply_borders(ws2)
        # 심각도 색상
        sev_col = list(violations_df.columns).index("심각도") + 1
        for r in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=r, column=sev_col)
            if cell.value == "치명":
                cell.fill = RED_FILL
            elif cell.value == "주의":
                cell.fill = YELLOW_FILL
    else:
        ws2.cell(row=1, column=1, value="하드필터 위반 케이스 없음")
    auto_width(ws2)

    # ── Sheet 3: Rank 변별력
    ws3 = wb.create_sheet("Rank_변별력_분석")
    # 요약 먼저
    ws3.cell(row=1, column=1, value="Rank별 평균 지표").font = Font(bold=True, size=11)
    ws3.cell(row=1, column=1).fill = SECTION_FILL
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    write_df(ws3, rank_summary_df, start_row=2)
    style_header(ws3, len(rank_summary_df.columns), row=2)
    for r in range(3, 3 + len(rank_summary_df)):
        for c in range(1, len(rank_summary_df.columns) + 1):
            ws3.cell(row=r, column=c).border = THIN_BORDER
            ws3.cell(row=r, column=c).font = CELL_FONT

    # 상세 (REF별)
    detail_start = 3 + len(rank_summary_df) + 2
    ws3.cell(row=detail_start, column=1, value="REF별 Rank 1~3 상세").font = Font(bold=True, size=11)
    ws3.cell(row=detail_start, column=1).fill = SECTION_FILL
    ws3.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=5)
    write_df(ws3, rank_detail_df, start_row=detail_start + 1)
    style_header(ws3, len(rank_detail_df.columns), row=detail_start + 1)
    for r in range(detail_start + 2, detail_start + 2 + len(rank_detail_df)):
        for c in range(1, len(rank_detail_df.columns) + 1):
            ws3.cell(row=r, column=c).border = THIN_BORDER
            ws3.cell(row=r, column=c).font = CELL_FONT
        # Rank 역전 표시
        inv_col = len(rank_detail_df.columns)
        cell = ws3.cell(row=r, column=inv_col)
        if cell.value == "Rank3>Rank1":
            cell.fill = RED_FILL
        elif cell.value == "차이없음":
            cell.fill = YELLOW_FILL
    # O/X 색상
    for r in range(detail_start + 2, detail_start + 2 + len(rank_detail_df)):
        for c in range(1, len(rank_detail_df.columns) + 1):
            cell = ws3.cell(row=r, column=c)
            if cell.value == "O":
                cell.fill = GREEN_FILL
            elif cell.value == "X":
                cell.fill = RED_FILL

    auto_width(ws3)

    # ── Sheet 4: 체계적 실패 패턴
    ws4 = wb.create_sheet("체계적_실패_패턴")
    write_df(ws4, patterns_df)
    style_header(ws4, len(patterns_df.columns))
    apply_borders(ws4)
    # 구분선 행 스타일
    for r in range(2, ws4.max_row + 1):
        if ws4.cell(row=r, column=1).value == "───":
            for c in range(1, ws4.max_column + 1):
                ws4.cell(row=r, column=c).fill = SEP_FILL
                ws4.cell(row=r, column=c).value = ""
    auto_width(ws4)

    # ── Sheet 5: 속성별 참고 통계
    ws5 = wb.create_sheet("속성별_참고_통계")
    ref_stats = build_reference_stats(df)
    write_df(ws5, ref_stats)
    style_header(ws5, len(ref_stats.columns))
    apply_borders(ws5)
    auto_width(ws5)

    wb.save(OUTPUT_PATH)
    print(f"  → 저장 완료: {OUTPUT_PATH}")

    # ── console summary ────────────────────────────────────────
    print("\n" + "=" * 60)
    print("주요 결과 요약")
    print("-" * 60)
    print(f"  총 {len(df)}건 / REF {df['REF_STYLE_CD'].nunique()}개 스타일")
    print()
    print(f"  하드필터 위반: {len(violations_df)}건")
    sev = violations_df["심각도"].value_counts() if len(violations_df) > 0 else {}
    if "치명" in sev:
        print(f"    치명(성별 위반): {sev['치명']}건")
    if "주의" in sev:
        print(f"    주의(가격 극단 등): {sev['주의']}건")

    print()
    print("  Rank 변별력:")
    for _, row in rank_summary_df.iterrows():
        if row[""] and row[""].startswith("Rank "):
            print(f"    {row['']}: DOMAIN {row['DOMAIN 일치율']}, "
                  f"상품명유사도 {row['평균 상품명유사도']}, "
                  f"가격차이 {row['평균 가격차이%']}")

    if len(rank_detail_df) > 0 and "Rank역전" in rank_detail_df.columns:
        inv = (rank_detail_df["Rank역전"] == "Rank3>Rank1").sum()
        nodiff = (rank_detail_df["Rank역전"] == "차이없음").sum()
        total_ref = len(rank_detail_df)
        print(f"    Rank역전(3>1): {inv}/{total_ref} ({inv/total_ref*100:.1f}%)")
        print(f"    Rank차이없음: {nodiff}/{total_ref} ({nodiff/total_ref*100:.1f}%)")

    # 과다추천 SIM
    sim_freq = df["SIMILAR_STYLE_CD"].value_counts()
    overused = sim_freq[sim_freq >= 10]
    if len(overused) > 0:
        print(f"\n  과다 추천 SIM: {len(overused)}개 스타일이 10회 이상 반환")
        for sim, cnt in overused.head(5).items():
            name = df[df["SIMILAR_STYLE_CD"] == sim].iloc[0]["SIM_PRDT_NM"]
            print(f"    {sim} ({name}): {cnt}회")

    print("\n" + "=" * 60)
    print("현업 리뷰 시트(Sheet 1)에서 판정(O/X/△) 입력 후")
    print("결과를 취합하면 모델 개선 방향을 구체화할 수 있습니다.")
    print("=" * 60)


if __name__ == "__main__":
    main()
